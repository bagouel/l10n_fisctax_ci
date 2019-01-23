# -*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *


class vat_declaration(osv.osv):
    _name="vat.declaration"
    _inherit="vat.declaration"

    def vat_report_xls(self, cr, uid, ids, context=None): # fonction report tva

        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\tva_template.xlsx"
        vat_dec=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path)
        wks = wbk.active

        thin1_border = Border(
            right=Side(style='thin'),  
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin2_border = Border(
            bottom=Side(style='thin')
            )
        thin8_border = Border(
            right=Side(style='thin'),
            bottom=Side(style='thin')
            )
        thin3_border = Border(
            left=Side(style='thin')
            )

        thin4_border = Border(
            bottom=Side(style='thin'),
            left=Side(style='thin')
            )
        thin5_border = Border(
            top=Side(style='thin'),
            right=Side(style='thin')
            )
        thin6_border = Border(
            right=Side(style='thin')
            )
        thin7_border = Border( 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        medium2_border = Border(
            right=Side(style='medium'), 
            )
        medium3_border = Border(
            right=Side(style='medium'),
            bottom=Side(style='thin') 
            )
        medium4_border = Border(
            left=Side(style='medium'),
            bottom=Side(style='thin') 
            )
        medium5_border = Border(
            left=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        medium6_border = Border(
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            )
        medium7_border = Border(
            left=Side(style='medium'),  
            top=Side(style='medium'), 
            )
        medium8_border = Border(
            right=Side(style='medium'),  
            bottom=Side(style='medium'), 
            )
        medium10_border = Border(
            bottom=Side(style='medium')
            )
        dashed_border = Border(
            bottom=Side(style='dashed')
            )


        wks['AE5'].border=thin3_border
        wks['A8'].border=thin3_border
        wks['A9'].border=thin3_border
        wks['A10'].border=thin4_border

        wks['AE5'].border=thin3_border
        wks['AE6'].border=thin4_border

        wks.merge_cells(start_row=4, start_column=1, end_row=6, end_column=9)
        wks['A4'].border=thin_border
        wks['B4'].border=thin_border
        wks['C4'].border=thin_border
        wks['D4'].border=thin_border
        wks['E4'].border=thin_border
        wks['F4'].border=thin_border
        wks['G4'].border=thin_border
        wks['H4'].border=thin_border
        wks['I4'].border=thin_border
        wks['A5'].border=thin_border
        wks['B5'].border=thin_border
        wks['C5'].border=thin_border
        wks['D5'].border=thin_border
        wks['E5'].border=thin_border
        wks['F5'].border=thin_border
        wks['G5'].border=thin_border
        wks['H5'].border=thin_border
        wks['I5'].border=thin_border
        wks['A6'].border=thin_border
        wks['B6'].border=thin_border
        wks['C6'].border=thin_border
        wks['D6'].border=thin_border
        wks['E6'].border=thin_border
        wks['F6'].border=thin_border
        wks['G6'].border=thin_border
        wks['H6'].border=thin_border
        wks['I6'].border=thin_border
        
        wks.merge_cells(start_row=4, start_column=31, end_row=6, end_column=41)
        wks['AE4'].border=thin_border
        wks['AF4'].border=thin_border
        wks['AG4'].border=thin_border
        wks['AH4'].border=thin_border
        wks['AI4'].border=thin_border
        wks['AJ4'].border=thin_border
        wks['AK4'].border=thin_border
        wks['AL4'].border=thin_border
        wks['AM4'].border=thin_border
        wks['AN4'].border=thin_border
        wks['AO4'].border=thin_border
        wks['AO4'].border=thin_border
        wks['AE5'].border=thin_border
        wks['AF5'].border=thin_border
        wks['AG5'].border=thin_border
        wks['AH5'].border=thin_border
        wks['AI5'].border=thin_border
        wks['AJ5'].border=thin_border
        wks['AK5'].border=thin_border
        wks['AL5'].border=thin_border
        wks['AM5'].border=thin_border
        wks['AN5'].border=thin_border
        wks['AO5'].border=thin_border
        wks['AO5'].border=thin_border
        wks['AE6'].border=thin_border
        wks['AF6'].border=thin_border
        wks['AG6'].border=thin_border
        wks['AH6'].border=thin_border
        wks['AI6'].border=thin_border
        wks['AJ6'].border=thin_border
        wks['AK6'].border=thin_border
        wks['AL6'].border=thin_border
        wks['AM6'].border=thin_border
        wks['AN6'].border=thin_border
        wks['AO6'].border=thin_border
        wks['AO6'].border=thin_border

        

        wks['AA18'].border=dashed_border
        wks['AB18'].border=dashed_border
        wks['AC18'].border=dashed_border
        wks['AD18'].border=dashed_border
        wks['AE18'].border=dashed_border
        wks['AF18'].border=dashed_border
        wks['AG18'].border=dashed_border
        wks['AH18'].border=dashed_border
        wks['AI18'].border=dashed_border
        wks['AJ18'].border=dashed_border
        wks['AK18'].border=dashed_border
        wks['AL18'].border=dashed_border
        wks['AM18'].border=dashed_border
        wks['AN18'].border=dashed_border
        wks['AO18'].border=dashed_border
        wks['AO18'].border=dashed_border

        wks.merge_cells('B20:Z20')
        wks['B20'].border=thin7_border
        wks['C20'].border=thin7_border
        wks['D20'].border=thin7_border
        wks['E20'].border=thin7_border
        wks['F20'].border=thin7_border
        wks['G20'].border=thin7_border
        wks['H20'].border=thin7_border
        wks['I20'].border=thin7_border
        wks['J20'].border=thin7_border
        wks['K20'].border=thin7_border
        wks['L20'].border=thin7_border
        wks['M20'].border=thin7_border
        wks['N20'].border=thin7_border
        wks['O20'].border=thin7_border
        wks['P20'].border=thin7_border
        wks['Q20'].border=thin7_border
        wks['R20'].border=thin7_border
        wks['S20'].border=thin7_border
        wks['T20'].border=thin7_border
        wks['U20'].border=thin7_border
        wks['V20'].border=thin7_border
        wks['W20'].border=thin7_border
        wks['X20'].border=thin7_border
        wks['Y20'].border=thin7_border
        wks['Z20'].border=thin7_border
        

        wks.merge_cells('A22:C22')
        wks['A22'].border=thin_border
        wks['B22'].border=thin_border
        wks['C22'].border=thin_border
        wks.merge_cells('A27:C27')
        wks['A27'].border=thin_border
        wks['B27'].border=thin_border
        wks['C27'].border=thin_border

        wks.merge_cells('A31:AA31')
        wks['A31'].border=thin_border
        wks['B31'].border=thin_border
        wks['C31'].border=thin_border
        wks['D31'].border=thin_border
        wks['E31'].border=thin_border
        wks['F31'].border=thin_border
        wks['G31'].border=thin_border
        wks['H31'].border=thin_border
        wks['I31'].border=thin_border
        wks['J31'].border=thin_border
        wks['K31'].border=thin_border
        wks['L31'].border=thin_border
        wks['M31'].border=thin_border
        wks['N31'].border=thin_border
        wks['O31'].border=thin_border
        wks['P31'].border=thin_border
        wks['Q31'].border=thin_border
        wks['R31'].border=thin_border
        wks['S31'].border=thin_border
        wks['T31'].border=thin_border
        wks['U31'].border=thin_border
        wks['V31'].border=thin_border
        wks['W31'].border=thin_border
        wks['X31'].border=thin_border
        wks['Y31'].border=thin_border
        wks['Z31'].border=thin_border
        wks['AA31'].border=thin_border
        

        wks.merge_cells('B32:Z32')
        wks['B32'].border=thin7_border
        wks['C32'].border=thin7_border
        wks['D32'].border=thin7_border
        wks['E32'].border=thin7_border
        wks['F32'].border=thin7_border
        wks['G32'].border=thin7_border
        wks['H32'].border=thin7_border
        wks['I32'].border=thin7_border
        wks['J32'].border=thin7_border
        wks['K32'].border=thin7_border
        wks['L32'].border=thin7_border
        wks['M32'].border=thin7_border
        wks['N32'].border=thin7_border
        wks['O32'].border=thin7_border
        wks['P32'].border=thin7_border
        wks['Q32'].border=thin7_border
        wks['R32'].border=thin7_border
        wks['S32'].border=thin7_border
        wks['T32'].border=thin7_border
        wks['U32'].border=thin7_border
        wks['V32'].border=thin7_border
        wks['W32'].border=thin7_border
        wks['X32'].border=thin7_border
        wks['Y32'].border=thin7_border
        wks['Z32'].border=thin7_border
        

        wks.merge_cells('C33:W33')
        wks['C33'].border=thin2_border
        wks['D33'].border=thin2_border
        wks['E33'].border=thin2_border
        wks['F33'].border=thin2_border
        wks['G33'].border=thin2_border
        wks['H33'].border=thin2_border
        wks['I33'].border=thin2_border
        wks['J33'].border=thin2_border
        wks['K33'].border=thin2_border
        wks['L33'].border=thin2_border
        wks['M33'].border=thin2_border
        wks['N33'].border=thin2_border
        wks['O33'].border=thin2_border
        wks['P33'].border=thin2_border
        wks['Q33'].border=thin2_border
        wks['R33'].border=thin2_border
        wks['S33'].border=thin2_border
        wks['T33'].border=thin2_border
        wks['U33'].border=thin2_border
        wks['V33'].border=thin2_border
        wks['W33'].border=thin2_border

        wks.merge_cells(start_row=34, start_column=25, end_row=35, end_column=27)
        wks['Y34'].border=thin_border
        wks['Z34'].border=thin_border
        wks['AA34'].border=thin_border
        wks['Y35'].border=thin_border
        wks['Z35'].border=thin_border
        wks['AA35'].border=thin_border

        wks.merge_cells(start_row=36, start_column=25, end_row=37, end_column=27)
        wks['Y36'].border=thin_border
        wks['Z36'].border=thin_border
        wks['AA36'].border=thin_border
        wks['Y37'].border=thin_border
        wks['Z37'].border=thin_border
        wks['AA37'].border=thin_border

        wks.merge_cells('Y33:AA33')
        wks['Y33'].border=thin_border
        wks['Z33'].border=thin_border
        wks['AA33'].border=thin_border
        

        wks.merge_cells('AB33:AK33')
        wks['AB33'].border=medium10_border
        wks['AC33'].border=medium10_border
        wks['AD33'].border=medium10_border
        wks['AE33'].border=medium10_border
        wks['AF33'].border=medium10_border
        wks['AG33'].border=medium10_border
        wks['AH33'].border=medium10_border
        wks['AI33'].border=medium10_border
        wks['AJ33'].border=medium10_border
        wks['AK33'].border=medium10_border

        wks.merge_cells(start_row=34, start_column=28, end_row=35, end_column=37)
        wks['AB34'].border=medium_border
        wks['AC34'].border=medium_border
        wks['AD34'].border=medium_border
        wks['AE34'].border=medium_border
        wks['AF34'].border=medium_border
        wks['AG34'].border=medium_border
        wks['AH34'].border=medium_border
        wks['AI34'].border=medium_border
        wks['AJ34'].border=medium_border
        wks['AK34'].border=medium_border
        wks['AB35'].border=medium_border
        wks['AC35'].border=medium_border
        wks['AD35'].border=medium_border
        wks['AE35'].border=medium_border
        wks['AF35'].border=medium_border
        wks['AG35'].border=medium_border
        wks['AH35'].border=medium_border
        wks['AI35'].border=medium_border
        wks['AJ35'].border=medium_border
        wks['AK35'].border=medium_border

        wks.merge_cells(start_row=36, start_column=28, end_row=37, end_column=37)
        wks['AB36'].border=medium_border
        wks['AC36'].border=medium_border
        wks['AD36'].border=medium_border
        wks['AE36'].border=medium_border
        wks['AF36'].border=medium_border
        wks['AG36'].border=medium_border
        wks['AH36'].border=medium_border
        wks['AI36'].border=medium_border
        wks['AJ36'].border=medium_border
        wks['AK36'].border=medium_border
        wks['AB37'].border=medium_border
        wks['AC37'].border=medium_border
        wks['AD37'].border=medium_border
        wks['AE37'].border=medium_border
        wks['AF37'].border=medium_border
        wks['AG37'].border=medium_border
        wks['AH37'].border=medium_border
        wks['AI37'].border=medium_border
        wks['AJ37'].border=medium_border
        wks['AK37'].border=medium_border
        

        wks['AC34'].border=dashed_border
        wks['AD34'].border=dashed_border
        wks['AE34'].border=dashed_border
        wks['AF34'].border=dashed_border
        wks['AG34'].border=dashed_border
        wks['AH34'].border=dashed_border
        wks['AI34'].border=dashed_border
        wks['AJ34'].border=dashed_border
        

        wks.merge_cells('B34:W34')
        wks['B34'].border=dashed_border
        wks['C34'].border=dashed_border
        wks['D34'].border=dashed_border
        wks['E34'].border=dashed_border
        wks['F34'].border=dashed_border
        wks['G34'].border=dashed_border
        wks['H34'].border=dashed_border
        wks['I34'].border=dashed_border
        wks['J34'].border=dashed_border
        wks['K34'].border=dashed_border
        wks['L34'].border=dashed_border
        wks['M34'].border=dashed_border
        wks['N34'].border=dashed_border
        wks['O34'].border=dashed_border
        wks['P34'].border=dashed_border
        wks['Q34'].border=dashed_border
        wks['R34'].border=dashed_border
        wks['S34'].border=dashed_border
        wks['T34'].border=dashed_border
        wks['U34'].border=dashed_border
        wks['V34'].border=dashed_border
        wks['W34'].border=dashed_border

        wks.merge_cells('B36:W36')
        wks['B36'].border=dashed_border
        wks['C36'].border=dashed_border
        wks['D36'].border=dashed_border
        wks['E36'].border=dashed_border
        wks['F36'].border=dashed_border
        wks['G36'].border=dashed_border
        wks['H36'].border=dashed_border
        wks['I36'].border=dashed_border
        wks['J36'].border=dashed_border
        wks['K36'].border=dashed_border
        wks['L36'].border=dashed_border
        wks['M36'].border=dashed_border
        wks['N36'].border=dashed_border
        wks['O36'].border=dashed_border
        wks['P36'].border=dashed_border
        wks['Q36'].border=dashed_border
        wks['R36'].border=dashed_border
        wks['S36'].border=dashed_border
        wks['T36'].border=dashed_border
        wks['U36'].border=dashed_border
        wks['V36'].border=dashed_border
        wks['W36'].border=dashed_border

        wks.merge_cells('B39:Z39')
        wks['B39'].border=thin7_border
        wks['C39'].border=thin7_border
        wks['D39'].border=thin7_border
        wks['E39'].border=thin7_border
        wks['F39'].border=thin7_border
        wks['G39'].border=thin7_border
        wks['H39'].border=thin7_border
        wks['I39'].border=thin7_border
        wks['J39'].border=thin7_border
        wks['K39'].border=thin7_border
        wks['L39'].border=thin7_border
        wks['M39'].border=thin7_border
        wks['N39'].border=thin7_border
        wks['O39'].border=thin7_border
        wks['P39'].border=thin7_border
        wks['Q39'].border=thin7_border
        wks['R39'].border=thin7_border
        wks['S39'].border=thin7_border
        wks['T39'].border=thin7_border
        wks['U39'].border=thin7_border
        wks['V39'].border=thin7_border
        wks['W39'].border=thin7_border
        wks['X39'].border=thin7_border
        wks['Y39'].border=thin7_border
        wks['Z39'].border=thin7_border

        wks.merge_cells('AB39:AK39')
        wks['AB39'].border=medium_border
        wks['AC39'].border=medium_border
        wks['AD39'].border=medium_border
        wks['AE39'].border=medium_border
        wks['AF39'].border=medium_border
        wks['AG39'].border=medium_border
        wks['AH39'].border=medium_border
        wks['AI39'].border=medium_border
        wks['AJ39'].border=medium_border
        wks['AK39'].border=medium_border
        
        
        

        wks.merge_cells('B41:Z41')
        wks['B41'].border=thin7_border
        wks['C41'].border=thin7_border
        wks['D41'].border=thin7_border
        wks['E41'].border=thin7_border
        wks['F41'].border=thin7_border
        wks['G41'].border=thin7_border
        wks['H41'].border=thin7_border
        wks['I41'].border=thin7_border
        wks['J41'].border=thin7_border
        wks['K41'].border=thin7_border
        wks['L41'].border=thin7_border
        wks['M41'].border=thin7_border
        wks['N41'].border=thin7_border
        wks['O41'].border=thin7_border
        wks['P41'].border=thin7_border
        wks['Q41'].border=thin7_border
        wks['R41'].border=thin7_border
        wks['S41'].border=thin7_border
        wks['T41'].border=thin7_border
        wks['U41'].border=thin7_border
        wks['V41'].border=thin7_border
        wks['W41'].border=thin7_border
        wks['X41'].border=thin7_border
        wks['Y41'].border=thin7_border
        wks['Z41'].border=thin7_border

        wks.merge_cells('AB41:AK41')
        wks['AB41'].border=medium_border
        wks['AC41'].border=medium_border
        wks['AD41'].border=medium_border
        wks['AE41'].border=medium_border
        wks['AF41'].border=medium_border
        wks['AG41'].border=medium_border
        wks['AH41'].border=medium_border
        wks['AI41'].border=medium_border
        wks['AJ41'].border=medium_border
        wks['AK41'].border=medium_border

        wks.merge_cells('P45:Z45')
        wks['P45'].border=dashed_border
        wks['Q45'].border=dashed_border
        wks['R45'].border=dashed_border
        wks['S45'].border=dashed_border
        wks['T45'].border=dashed_border
        wks['U45'].border=dashed_border
        wks['V45'].border=dashed_border
        wks['W45'].border=dashed_border
        wks['X45'].border=dashed_border
        wks['Y45'].border=dashed_border
        wks['Z45'].border=dashed_border

        wks.merge_cells('P47:Z47')
        wks['P47'].border=dashed_border
        wks['Q47'].border=dashed_border
        wks['R47'].border=dashed_border
        wks['S47'].border=dashed_border
        wks['T47'].border=dashed_border
        wks['U47'].border=dashed_border
        wks['V47'].border=dashed_border
        wks['W47'].border=dashed_border
        wks['X47'].border=dashed_border
        wks['Y47'].border=dashed_border
        wks['Z47'].border=dashed_border
        
        wks.merge_cells('O49:Z49')
        wks['O49'].border=thin2_border
        wks['P49'].border=thin2_border
        wks['Q49'].border=thin2_border
        wks['R49'].border=thin2_border
        wks['S49'].border=thin2_border
        wks['T49'].border=thin2_border
        wks['U49'].border=thin2_border
        wks['V49'].border=thin2_border
        wks['W49'].border=thin2_border
        wks['X49'].border=thin2_border
        wks['Y49'].border=thin2_border
        wks['Z49'].border=thin2_border
        


        wks.merge_cells('AB49:AK49')
        wks['AB49'].border=medium_border
        wks['AC49'].border=medium_border
        wks['AD49'].border=medium_border
        wks['AE49'].border=medium_border
        wks['AF49'].border=medium_border
        wks['AG49'].border=medium_border
        wks['AH49'].border=medium_border
        wks['AI49'].border=medium_border
        wks['AJ49'].border=medium_border
        wks['AK49'].border=medium_border

        wks.merge_cells('B43:Z43')
        wks['B43'].border=thin7_border
        wks['C43'].border=thin7_border
        wks['D43'].border=thin7_border
        wks['E43'].border=thin7_border
        wks['F43'].border=thin7_border
        wks['G43'].border=thin7_border
        wks['H43'].border=thin7_border
        wks['I43'].border=thin7_border
        wks['J43'].border=thin7_border
        wks['K43'].border=thin7_border
        wks['L43'].border=thin7_border
        wks['M43'].border=thin7_border
        wks['N43'].border=thin7_border
        wks['O43'].border=thin7_border
        wks['P43'].border=thin7_border
        wks['Q43'].border=thin7_border
        wks['R43'].border=thin7_border
        wks['S43'].border=thin7_border
        wks['T43'].border=thin7_border
        wks['U43'].border=thin7_border
        wks['V43'].border=thin7_border
        wks['W43'].border=thin7_border
        wks['X43'].border=thin7_border
        wks['Y43'].border=thin7_border
        wks['Z43'].border=thin7_border
       
        

        wks.merge_cells('P21:Z21')
        wks['P21'].border=dashed_border
        wks['Q21'].border=dashed_border
        wks['R21'].border=dashed_border
        wks['S21'].border=dashed_border
        wks['T21'].border=dashed_border
        wks['U21'].border=dashed_border
        wks['V21'].border=dashed_border
        wks['W21'].border=dashed_border
        wks['X21'].border=dashed_border
        wks['Y21'].border=dashed_border
        wks['Z21'].border=dashed_border
        wks.merge_cells('P22:Z22')
        wks['P22'].border=dashed_border
        wks['Q22'].border=dashed_border
        wks['R22'].border=dashed_border
        wks['S22'].border=dashed_border
        wks['T22'].border=dashed_border
        wks['U22'].border=dashed_border
        wks['V22'].border=dashed_border
        wks['W22'].border=dashed_border
        wks['X22'].border=dashed_border
        wks['Y22'].border=dashed_border
        wks['Z22'].border=dashed_border
        wks.merge_cells('P23:Z23')
        wks['P23'].border=dashed_border
        wks['Q23'].border=dashed_border
        wks['R23'].border=dashed_border
        wks['S23'].border=dashed_border
        wks['T23'].border=dashed_border
        wks['U23'].border=dashed_border
        wks['V23'].border=dashed_border
        wks['W23'].border=dashed_border
        wks['X23'].border=dashed_border
        wks['Y23'].border=dashed_border
        wks['Z23'].border=dashed_border
        wks.merge_cells('P24:Z24')
        wks['P24'].border=dashed_border
        wks['Q24'].border=dashed_border
        wks['R24'].border=dashed_border
        wks['S24'].border=dashed_border
        wks['T24'].border=dashed_border
        wks['U24'].border=dashed_border
        wks['V24'].border=dashed_border
        wks['W24'].border=dashed_border
        wks['X24'].border=dashed_border
        wks['Y24'].border=dashed_border
        wks['Z24'].border=dashed_border
        wks.merge_cells('P25:Z25')
        wks['P25'].border=dashed_border
        wks['Q25'].border=dashed_border
        wks['R25'].border=dashed_border
        wks['S25'].border=dashed_border
        wks['T25'].border=dashed_border
        wks['U25'].border=dashed_border
        wks['V25'].border=dashed_border
        wks['W25'].border=dashed_border
        wks['X25'].border=dashed_border
        wks['Y25'].border=dashed_border
        wks['Z25'].border=dashed_border
        wks.merge_cells('P28:Z28')
        wks['P28'].border=dashed_border
        wks['Q28'].border=dashed_border
        wks['R28'].border=dashed_border
        wks['S28'].border=dashed_border
        wks['T28'].border=dashed_border
        wks['U28'].border=dashed_border
        wks['V28'].border=dashed_border
        wks['W28'].border=dashed_border
        wks['X28'].border=dashed_border
        wks['Y28'].border=dashed_border
        wks['Z28'].border=dashed_border
        wks.merge_cells('P29:Z29')
        wks['P29'].border=dashed_border
        wks['Q29'].border=dashed_border
        wks['R29'].border=dashed_border
        wks['S29'].border=dashed_border
        wks['T29'].border=dashed_border
        wks['U29'].border=dashed_border
        wks['V29'].border=dashed_border
        wks['W29'].border=dashed_border
        wks['X29'].border=dashed_border
        wks['Y29'].border=dashed_border
        wks['Z29'].border=dashed_border
        
        wks['AA17']=vat_dec.tax_service
        wks['AE7']=vat_dec.date
        wks['P21']=vat_dec.revenu_wout_vat
        wks['P2']=vat_dec.export_deduction
        wks['P23']=vat_dec.legal_ops_exempt_vat_revenu
        wks['P24']=vat_dec.conv_ops_exempt_vat_revenu
        wks['P25']=vat_dec.other_non_vat_revenu
        wks['P26']=vat_dec.difference
        wks['P27']=vat_dec.vat_delivery_onself_revenu
        wks['P28']=vat_dec.taxable_revenu_wout_vat
        wks['B34']=vat_dec.normal_rate
        wks['B36']=vat_dec.minimal_rate
        wks['Y34']=vat_dec.normal_rate_vat_amount
        wks['Y36']=vat_dec.minimal_rate_vat_amount
        wks['AB34']=vat_dec.normal_rate_revenu_amount
        wks['AB36']=vat_dec.minimal_rate_revenu_amount
        wks['AB41']=vat_dec.vat_gross_total
        wks['AB49']=vat_dec.deductible_vat_total
        wks['AB52']=vat_dec.credit_vat_toreport
        wks['AB53']=vat_dec.credit_vat_torefund
        wks['P45']=vat_dec.monthly_deductible_vat
        wks['P47']=vat_dec.lastest_vat_credit

        wks.merge_cells('A51:AA51')
        wks['A51'].border=medium_border
        wks['B51'].border=medium_border
        wks['C51'].border=medium_border
        wks['D51'].border=medium_border
        wks['E51'].border=medium_border
        wks['F51'].border=medium_border
        wks['G51'].border=medium_border
        wks['H51'].border=medium_border
        wks['I51'].border=medium_border
        wks['J51'].border=medium_border
        wks['K51'].border=medium_border
        wks['L51'].border=medium_border
        wks['M51'].border=medium_border
        wks['N51'].border=medium_border
        wks['O51'].border=medium_border
        wks['P51'].border=medium_border
        wks['Q51'].border=medium_border
        wks['R51'].border=medium_border
        wks['S51'].border=medium_border
        wks['T51'].border=medium_border
        wks['U51'].border=medium_border
        wks['V51'].border=medium_border
        wks['W51'].border=medium_border
        wks['X51'].border=medium_border
        wks['Y51'].border=medium_border
        wks['Z51'].border=medium_border
        wks['AA51'].border=medium_border

        wks.merge_cells('AB51:AK51')
        wks['AB51'].border=medium_border
        wks['AC51'].border=medium_border
        wks['AD51'].border=medium_border
        wks['AE51'].border=medium_border
        wks['AF51'].border=medium_border
        wks['AG51'].border=medium_border
        wks['AH51'].border=medium_border
        wks['AI51'].border=medium_border
        wks['AJ51'].border=medium_border
        wks['AK51'].border=medium_border

        wks['A53'].border=thin2_border
        wks['B53'].border=thin2_border
        wks['C53'].border=thin2_border
        wks['D53'].border=thin2_border
        wks['E53'].border=thin2_border
        wks['F53'].border=thin2_border
        wks['G53'].border=thin2_border
        wks['H53'].border=thin2_border
        wks['I53'].border=thin2_border
        wks['J53'].border=thin2_border
        wks['K53'].border=thin2_border
        wks['L53'].border=thin2_border
        wks['M53'].border=thin2_border
        wks['N53'].border=thin2_border
        wks['O53'].border=thin2_border
        wks['P53'].border=thin2_border
        wks['Q53'].border=thin2_border
        wks['R53'].border=thin2_border
        wks['S53'].border=thin2_border
        wks['T53'].border=thin2_border
        wks['U53'].border=thin2_border
        wks['V53'].border=thin2_border
        wks['W53'].border=thin2_border
        wks['X53'].border=thin2_border
        wks['Y53'].border=thin2_border
        wks['Z53'].border=thin2_border
       

        wks.merge_cells(start_row=52, start_column=28, end_row=53, end_column=37)
        wks['AB52'].border=medium_border
        wks['AC52'].border=medium_border
        wks['AD52'].border=medium_border
        wks['AE52'].border=medium_border
        wks['AF52'].border=medium_border
        wks['AG52'].border=medium_border
        wks['AH52'].border=medium_border
        wks['AI52'].border=medium_border
        wks['AJ52'].border=medium_border
        wks['AK52'].border=medium_border
        wks['AB53'].border=medium_border
        wks['AC53'].border=medium_border
        wks['AD53'].border=medium_border
        wks['AE53'].border=medium_border
        wks['AF53'].border=medium_border
        wks['AG53'].border=medium_border
        wks['AH53'].border=medium_border
        wks['AI53'].border=medium_border
        wks['AJ53'].border=medium_border
        wks['AK53'].border=medium_border
       

        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'vat_report_xls','vat_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'vat.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }