# -*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *


class wagetax_declaration(osv.osv):
    _name="wagetax.declaration"
    _inherit="wagetax.declaration"


    def wagetax_report_xls(self, cr, uid, ids, context=None): # fonction report its

        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\its_template.xlsx"
        wagetax_dec=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path)
        wks = wbk.active

        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )

        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        medium1_border =  Border( 
            right=Side(style='medium'), 
            )
        medium2_border =  Border( 
            top=Side(style='medium'), 
            bottom=Side(style='dashed') 
            )
        
        dashed_border = Border(
            bottom=Side(style='dashed')
            )

        wks['D1']=wagetax_dec.company_tax_code
        wks.merge_cells(start_row=2, start_column=2, end_row=3, end_column=10)
        wks['B2'].border=thin_border
        wks['C2'].border=thin_border
        wks['D2'].border=thin_border
        wks['E2'].border=thin_border
        wks['D2'].border=thin_border
        wks['E2'].border=thin_border
        wks['F2'].border=thin_border
        wks['G2'].border=thin_border
        wks['H2'].border=thin_border
        wks['B3'].border=thin_border
        wks['C3'].border=thin_border
        wks['D3'].border=thin_border
        wks['E3'].border=thin_border
        wks['D3'].border=thin_border
        wks['E3'].border=thin_border
        wks['F3'].border=thin_border
        wks['G3'].border=thin_border
        wks['H3'].border=thin_border

        wks.merge_cells(start_row=4, start_column=2, end_row=9, end_column=10)
        wks['B4'].border=thin_border
        wks['C4'].border=thin_border
        wks['D4'].border=thin_border
        wks['E4'].border=thin_border
        wks['D4'].border=thin_border
        wks['E4'].border=thin_border
        wks['F4'].border=thin_border
        wks['G4'].border=thin_border
        wks['H4'].border=thin_border
        wks['B5'].border=thin_border
        wks['C5'].border=thin_border
        wks['D5'].border=thin_border
        wks['E5'].border=thin_border
        wks['D5'].border=thin_border
        wks['E5'].border=thin_border
        wks['F5'].border=thin_border
        wks['G5'].border=thin_border
        wks['H5'].border=thin_border
        wks['B6'].border=thin_border
        wks['C6'].border=thin_border
        wks['D6'].border=thin_border
        wks['E6'].border=thin_border
        wks['D6'].border=thin_border
        wks['E6'].border=thin_border
        wks['F6'].border=thin_border
        wks['G6'].border=thin_border
        wks['H6'].border=thin_border
        wks['B7'].border=thin_border
        wks['C7'].border=thin_border
        wks['D7'].border=thin_border
        wks['E7'].border=thin_border
        wks['D7'].border=thin_border
        wks['E7'].border=thin_border
        wks['F7'].border=thin_border
        wks['G7'].border=thin_border
        wks['H7'].border=thin_border
        wks['B8'].border=thin_border
        wks['C8'].border=thin_border
        wks['D8'].border=thin_border
        wks['E8'].border=thin_border
        wks['D8'].border=thin_border
        wks['E8'].border=thin_border
        wks['F8'].border=thin_border
        wks['G8'].border=thin_border
        wks['H8'].border=thin_border
        wks['B9'].border=thin_border
        wks['C9'].border=thin_border
        wks['D9'].border=thin_border
        wks['E9'].border=thin_border
        wks['D9'].border=thin_border
        wks['E9'].border=thin_border
        wks['F9'].border=thin_border
        wks['G9'].border=thin_border
        wks['H9'].border=thin_border

        wks.merge_cells(start_row=2, start_column=35, end_row=3, end_column=41)
        wks['AI2'].border=thin_border
        wks['AJ2'].border=thin_border
        wks['AK2'].border=thin_border
        wks['AM2'].border=thin_border
        wks['AO2'].border=thin_border
        wks['AP2'].border=thin_border
        wks['AQ2'].border=thin_border
        wks['AI3'].border=thin_border
        wks['AJ3'].border=thin_border
        wks['AK3'].border=thin_border
        wks['AM3'].border=thin_border
        wks['AO3'].border=thin_border
        wks['AP3'].border=thin_border
        wks['AQ3'].border=thin_border

        wks.merge_cells(start_row=4, start_column=35, end_row=9, end_column=41)
        wks['AI4'].border=thin_border
        wks['AJ4'].border=thin_border
        wks['AK4'].border=thin_border
        wks['AM4'].border=thin_border
        wks['AO4'].border=thin_border
        wks['AP4'].border=thin_border
        wks['AQ4'].border=thin_border
        wks['AI5'].border=thin_border
        wks['AJ5'].border=thin_border
        wks['AK5'].border=thin_border
        wks['AM5'].border=thin_border
        wks['AO5'].border=thin_border
        wks['AP5'].border=thin_border
        wks['AQ5'].border=thin_border
        wks['AI6'].border=thin_border
        wks['AJ6'].border=thin_border
        wks['AK6'].border=thin_border
        wks['AM6'].border=thin_border
        wks['AO6'].border=thin_border
        wks['AP6'].border=thin_border
        wks['AQ6'].border=thin_border
        wks['AI7'].border=thin_border
        wks['AJ7'].border=thin_border
        wks['AK7'].border=thin_border
        wks['AM7'].border=thin_border
        wks['AO7'].border=thin_border
        wks['AP7'].border=thin_border
        wks['AQ7'].border=thin_border
        wks['AI8'].border=thin_border
        wks['AJ8'].border=thin_border
        wks['AK8'].border=thin_border
        wks['AM8'].border=thin_border
        wks['AO8'].border=thin_border
        wks['AP8'].border=thin_border
        wks['AQ8'].border=thin_border
        wks['AI9'].border=thin_border
        wks['AJ9'].border=thin_border
        wks['AK9'].border=thin_border
        wks['AM9'].border=thin_border
        wks['AO9'].border=thin_border
        wks['AP9'].border=thin_border
        wks['AQ9'].border=thin_border

        wks.merge_cells('AH16:AV16')
        wks['AH16']=wagetax_dec.tax_service
        wks['AH16'].border=dashed_border
        wks['AI16'].border=dashed_border
        wks['AJ16'].border=dashed_border
        wks['AK16'].border=dashed_border
        wks['AL16'].border=dashed_border
        wks['AM16'].border=dashed_border
        wks['AN16'].border=dashed_border
        wks['AO16'].border=dashed_border
        wks['AP16'].border=dashed_border
        wks['AQ16'].border=dashed_border
        wks['AR16'].border=dashed_border
        wks['AS16'].border=dashed_border
        wks['AT16'].border=dashed_border
        wks['AU16'].border=dashed_border
        wks['AV16'].border=dashed_border
        

        wks.merge_cells('M20:AW20')
        wks['M20']=wagetax_dec.company
        wks['M20'].border=medium2_border
        wks['N20'].border=medium2_border
        wks['O20'].border=medium2_border
        wks['P20'].border=medium2_border
        wks['Q20'].border=medium2_border
        wks['R20'].border=medium2_border
        wks['S20'].border=medium2_border
        wks['T20'].border=medium2_border
        wks['U20'].border=medium2_border
        wks['V20'].border=medium2_border
        wks['W20'].border=medium2_border
        wks['X20'].border=medium2_border
        wks['Y20'].border=medium2_border
        wks['Z20'].border=medium2_border
        wks['AA20'].border=medium2_border
        wks['AB20'].border=medium2_border
        wks['AC20'].border=medium2_border
        wks['AD20'].border=medium2_border
        wks['AE20'].border=medium2_border
        wks['AF20'].border=medium2_border
        wks['AG20'].border=medium2_border
        wks['AH20'].border=medium2_border
        wks['AI20'].border=medium2_border
        wks['AJ20'].border=medium2_border
        wks['AK20'].border=medium2_border
        wks['AL20'].border=medium2_border
        wks['AM20'].border=medium2_border
        wks['AN20'].border=medium2_border
        wks['AO20'].border=medium2_border
        wks['AP20'].border=medium2_border
        wks['AQ20'].border=medium2_border
        wks['AR20'].border=medium2_border
        wks['AS20'].border=medium2_border
        wks['AT20'].border=medium2_border
        wks['AU20'].border=medium2_border
        wks['AV20'].border=medium2_border
        wks['AW20'].border=medium2_border

        wks.merge_cells('D23:X23')
        wks['D23']=wagetax_dec.initials
        wks['D23'].border=dashed_border
        wks['E23'].border=dashed_border
        wks['F23'].border=dashed_border
        wks['G23'].border=dashed_border
        wks['H23'].border=dashed_border
        wks['I23'].border=dashed_border
        wks['J23'].border=dashed_border
        wks['K23'].border=dashed_border
        wks['L23'].border=dashed_border
        wks['M23'].border=dashed_border
        wks['N23'].border=dashed_border
        wks['O23'].border=dashed_border
        wks['P23'].border=dashed_border
        wks['Q23'].border=dashed_border
        wks['R23'].border=dashed_border
        wks['S23'].border=dashed_border
        wks['T23'].border=dashed_border
        wks['U23'].border=dashed_border
        wks['V23'].border=dashed_border
        wks['W23'].border=dashed_border
        wks['X23'].border=dashed_border

        wks['AX23'].border=medium1_border


        wks.merge_cells('G24:AW24')
        wks['G24']=wagetax_dec.company_goal
        wks['G24'].border=dashed_border
        wks['H24'].border=dashed_border
        wks['I24'].border=dashed_border
        wks['J24'].border=dashed_border
        wks['K24'].border=dashed_border
        wks['L24'].border=dashed_border
        wks['M24'].border=dashed_border
        wks['N24'].border=dashed_border
        wks['O24'].border=dashed_border
        wks['P24'].border=dashed_border
        wks['Q24'].border=dashed_border
        wks['R24'].border=dashed_border
        wks['S24'].border=dashed_border
        wks['T24'].border=dashed_border
        wks['U24'].border=dashed_border
        wks['V24'].border=dashed_border
        wks['W24'].border=dashed_border
        wks['X24'].border=dashed_border
        wks['Y24'].border=dashed_border
        wks['Z24'].border=dashed_border
        wks['AA24'].border=dashed_border
        wks['AB24'].border=dashed_border
        wks['AC24'].border=dashed_border
        wks['AD24'].border=dashed_border
        wks['AE24'].border=dashed_border
        wks['AF24'].border=dashed_border
        wks['AG24'].border=dashed_border
        wks['AH24'].border=dashed_border
        wks['AI24'].border=dashed_border
        wks['AJ24'].border=dashed_border
        wks['AK24'].border=dashed_border
        wks['AL24'].border=dashed_border
        wks['AM24'].border=dashed_border
        wks['AN24'].border=dashed_border
        wks['AO24'].border=dashed_border
        wks['AP24'].border=dashed_border
        wks['AQ24'].border=dashed_border
        wks['AR24'].border=dashed_border
        wks['AS24'].border=dashed_border
        wks['AT24'].border=dashed_border
        wks['AU24'].border=dashed_border
        wks['AV24'].border=dashed_border
        wks['AW24'].border=dashed_border

        wks.merge_cells('H25:Z25')
        wks['H25']=wagetax_dec.street
        wks['H25'].border=dashed_border
        wks['I25'].border=dashed_border
        wks['J25'].border=dashed_border
        wks['K25'].border=dashed_border
        wks['L25'].border=dashed_border
        wks['M25'].border=dashed_border
        wks['N25'].border=dashed_border
        wks['O25'].border=dashed_border
        wks['P25'].border=dashed_border
        wks['Q25'].border=dashed_border
        wks['R25'].border=dashed_border
        wks['S25'].border=dashed_border
        wks['T25'].border=dashed_border
        wks['U25'].border=dashed_border
        wks['V25'].border=dashed_border
        wks['W25'].border=dashed_border
        wks['X25'].border=dashed_border
        wks['Y25'].border=dashed_border
        wks['Z25'].border=dashed_border
       
        wks.merge_cells('AB25:AO25')
        wks['AB25']=wagetax_dec.pobox
        wks['AB25'].border=dashed_border
        wks['AC25'].border=dashed_border
        wks['AD25'].border=dashed_border
        wks['AE25'].border=dashed_border
        wks['AF25'].border=dashed_border
        wks['AG25'].border=dashed_border
        wks['AH25'].border=dashed_border
        wks['AI25'].border=dashed_border
        wks['AJ25'].border=dashed_border
        wks['AK25'].border=dashed_border
        wks['AL25'].border=dashed_border
        wks['AM25'].border=dashed_border
        wks['AN25'].border=dashed_border
        wks['AO25'].border=dashed_border

        wks.merge_cells('AQ25:AW25')
        wks['AQ25']=wagetax_dec.phone
        wks['AQ25'].border=dashed_border
        wks['AR25'].border=dashed_border
        wks['AS25'].border=dashed_border
        wks['AT25'].border=dashed_border
        wks['AU25'].border=dashed_border
        wks['AV25'].border=dashed_border
        wks['AW25'].border=dashed_border


        wks.merge_cells('G26:AD26')
        wks['G26']=wagetax_dec.district
        wks['G26'].border=dashed_border
        wks['H26'].border=dashed_border
        wks['I26'].border=dashed_border
        wks['J26'].border=dashed_border
        wks['K26'].border=dashed_border
        wks['L26'].border=dashed_border
        wks['M26'].border=dashed_border
        wks['N26'].border=dashed_border
        wks['O26'].border=dashed_border
        wks['P26'].border=dashed_border
        wks['Q26'].border=dashed_border
        wks['R26'].border=dashed_border
        wks['S26'].border=dashed_border
        wks['T26'].border=dashed_border
        wks['U26'].border=dashed_border
        wks['V26'].border=dashed_border
        wks['W26'].border=dashed_border
        wks['X26'].border=dashed_border
        wks['Y26'].border=dashed_border
        wks['Z26'].border=dashed_border
        wks['AA26'].border=dashed_border
        wks['AB26'].border=dashed_border
        wks['AC26'].border=dashed_border
        wks['AD26'].border=dashed_border


        wks.merge_cells('AG26:AW26')
        wks['AG26']=wagetax_dec.street2
        wks['AG26'].border=dashed_border
        wks['AH26'].border=dashed_border
        wks['AI26'].border=dashed_border
        wks['AJ26'].border=dashed_border
        wks['AK26'].border=dashed_border
        wks['AL26'].border=dashed_border
        wks['AM26'].border=dashed_border
        wks['AN26'].border=dashed_border
        wks['AO26'].border=dashed_border
        wks['AP26'].border=dashed_border
        wks['AQ26'].border=dashed_border
        wks['AR26'].border=dashed_border
        wks['AS26'].border=dashed_border
        wks['AT26'].border=dashed_border
        wks['AU26'].border=dashed_border
        wks['AV26'].border=dashed_border
        wks['AW26'].border=dashed_border

        wks.merge_cells('J27:AW27')
        wks['J27']=wagetax_dec.email
        wks['J27'].border=dashed_border
        wks['K27'].border=dashed_border
        wks['L27'].border=dashed_border
        wks['M27'].border=dashed_border
        wks['N27'].border=dashed_border
        wks['O27'].border=dashed_border
        wks['P27'].border=dashed_border
        wks['Q27'].border=dashed_border
        wks['R27'].border=dashed_border
        wks['S27'].border=dashed_border
        wks['T27'].border=dashed_border
        wks['U27'].border=dashed_border
        wks['V27'].border=dashed_border
        wks['W27'].border=dashed_border
        wks['X27'].border=dashed_border
        wks['Y27'].border=dashed_border
        wks['Z27'].border=dashed_border
        wks['AA27'].border=dashed_border
        wks['AB27'].border=dashed_border
        wks['AC27'].border=dashed_border
        wks['AD27'].border=dashed_border
        wks['AE27'].border=dashed_border
        wks['AF27'].border=dashed_border
        wks['AG27'].border=dashed_border
        wks['AH27'].border=dashed_border
        wks['AI27'].border=dashed_border
        wks['AJ27'].border=dashed_border
        wks['AK27'].border=dashed_border
        wks['AL27'].border=dashed_border
        wks['AM27'].border=dashed_border
        wks['AN27'].border=dashed_border
        wks['AO27'].border=dashed_border
        wks['AP27'].border=dashed_border
        wks['AQ27'].border=dashed_border
        wks['AR27'].border=dashed_border
        wks['AS27'].border=dashed_border
        wks['AT27'].border=dashed_border
        wks['AU27'].border=dashed_border
        wks['AV27'].border=dashed_border
        wks['AW27'].border=dashed_border


        wks.merge_cells('B50:C50')
        wks['B50'].border=thin_border
        wks['C50'].border=thin_border

        wks.merge_cells('D50:O50')
        wks['D50'].border=thin_border
        wks['E50'].border=thin_border
        wks['F50'].border=thin_border
        wks['G50'].border=thin_border
        wks['H50'].border=thin_border
        wks['I50'].border=thin_border
        wks['J50'].border=thin_border
        wks['K50'].border=thin_border
        wks['L50'].border=thin_border
        wks['M50'].border=thin_border
        wks['N50'].border=thin_border
        wks['O50'].border=thin_border

        wks.merge_cells('P50:T50')
        wks['P50'].border=thin_border
        wks['Q50'].border=thin_border
        wks['R50'].border=thin_border
        wks['S50'].border=thin_border
        wks['T50'].border=thin_border

        wks.merge_cells('U50:V50')
        wks['U50'].border=thin_border
        wks['V50'].border=thin_border

        wks.merge_cells('W50:AF50')
        wks['W50'].border=thin_border
        wks['X50'].border=thin_border
        wks['Y50'].border=thin_border
        wks['Z50'].border=thin_border
        wks['AA50'].border=thin_border
        wks['AB50'].border=thin_border
        wks['AC50'].border=thin_border
        wks['AD50'].border=thin_border
        wks['AE50'].border=thin_border
        wks['AF50'].border=thin_border

        wks.merge_cells('AG50:AK50')
        wks['AG50'].border=thin_border
        wks['AH50'].border=thin_border
        wks['AI50'].border=thin_border
        wks['AJ50'].border=thin_border
        wks['AK50'].border=thin_border

        wks.merge_cells('AL50:AQ50')
        wks['AL50'].border=thin_border
        wks['AM50'].border=thin_border
        wks['AN50'].border=thin_border
        wks['AO50'].border=thin_border
        wks['AP50'].border=thin_border
        wks['AQ50'].border=thin_border

        wks.merge_cells('AR50:AX50')
        wks['AR50'].border=thin_border
        wks['AS50'].border=thin_border
        wks['AT50'].border=thin_border
        wks['AU50'].border=thin_border
        wks['AV50'].border=thin_border
        wks['AW50'].border=thin_border
        wks['AX50'].border=thin_border



        wks.merge_cells('B51:C51')
        wks['B51'].border=thin_border
        wks['C51'].border=thin_border

        wks.merge_cells('D51:O51')
        wks['D51'].border=thin_border
        wks['E51'].border=thin_border
        wks['F51'].border=thin_border
        wks['G51'].border=thin_border
        wks['H51'].border=thin_border
        wks['I51'].border=thin_border
        wks['J51'].border=thin_border
        wks['K51'].border=thin_border
        wks['L51'].border=thin_border
        wks['M51'].border=thin_border
        wks['N51'].border=thin_border
        wks['O51'].border=thin_border

        wks.merge_cells('P51:T51')
        wks['P51']=wagetax_dec.amount_tv1
        wks['P51'].border=thin_border
        wks['Q51'].border=thin_border
        wks['R51'].border=thin_border
        wks['S51'].border=thin_border
        wks['T51'].border=thin_border

        wks.merge_cells('U51:V51')
        wks['U51']=wagetax_dec.workforce_1
        wks['U51'].border=thin_border
        wks['V51'].border=thin_border

        wks.merge_cells('W51:AF51')
        wks['W51']=wagetax_dec.amount_exo1
        wks['W51'].border=thin_border
        wks['X51'].border=thin_border
        wks['Y51'].border=thin_border
        wks['Z51'].border=thin_border
        wks['AA51'].border=thin_border
        wks['AB51'].border=thin_border
        wks['AC51'].border=thin_border
        wks['AD51'].border=thin_border
        wks['AE51'].border=thin_border
        wks['AF51'].border=thin_border

        wks.merge_cells('AG51:AK51')
        wks['AG51']=wagetax_dec.amount_tax1
        wks['AG51'].border=thin_border
        wks['AH51'].border=thin_border
        wks['AI51'].border=thin_border
        wks['AJ51'].border=thin_border
        wks['AK51'].border=thin_border

        wks.merge_cells('AL51:AQ51')
        wks['AL51']=wagetax_dec.taxreduction_1
        wks['AL51'].border=thin_border
        wks['AM51'].border=thin_border
        wks['AN51'].border=thin_border
        wks['AO51'].border=thin_border
        wks['AP51'].border=thin_border
        wks['AQ51'].border=thin_border

        wks.merge_cells('AR51:AX51')
        wks['AR51']=wagetax_dec.revenu_ni1
        wks['AR51'].border=thin_border
        wks['AS51'].border=thin_border
        wks['AT51'].border=thin_border
        wks['AU51'].border=thin_border
        wks['AV51'].border=thin_border
        wks['AW51'].border=thin_border
        wks['AX51'].border=thin_border


        wks.merge_cells('B52:C52')
        wks['B52'].border=thin_border
        wks['C52'].border=thin_border

        wks.merge_cells('D52:O52')
        wks['D52'].border=thin_border
        wks['E52'].border=thin_border
        wks['F52'].border=thin_border
        wks['G52'].border=thin_border
        wks['H52'].border=thin_border
        wks['I52'].border=thin_border
        wks['J52'].border=thin_border
        wks['K52'].border=thin_border
        wks['L52'].border=thin_border
        wks['M52'].border=thin_border
        wks['N52'].border=thin_border
        wks['O52'].border=thin_border

        wks.merge_cells('P52:T52')
        wks['P52']=wagetax_dec.amount_tv2
        wks['P52'].border=thin_border
        wks['Q52'].border=thin_border
        wks['R52'].border=thin_border
        wks['S52'].border=thin_border
        wks['T52'].border=thin_border

        wks.merge_cells('U52:V52')
        wks['U52']=wagetax_dec.workforce_2
        wks['U52'].border=thin_border
        wks['V52'].border=thin_border

        wks.merge_cells('W52:AF52')
        wks['W52']=wagetax_dec.amount_exo2
        wks['W52'].border=thin_border
        wks['X52'].border=thin_border
        wks['Y52'].border=thin_border
        wks['Z52'].border=thin_border
        wks['AA52'].border=thin_border
        wks['AB52'].border=thin_border
        wks['AC52'].border=thin_border
        wks['AD52'].border=thin_border
        wks['AE52'].border=thin_border
        wks['AF52'].border=thin_border

        wks.merge_cells('AG52:AK52')
        wks['AG52']=wagetax_dec.amount_tax2
        wks['AG52'].border=thin_border
        wks['AH52'].border=thin_border
        wks['AI52'].border=thin_border
        wks['AJ52'].border=thin_border
        wks['AK52'].border=thin_border

        wks.merge_cells('AL52:AQ52')
        wks['AL52']=wagetax_dec.taxreduction_2
        wks['AL52'].border=thin_border
        wks['AM52'].border=thin_border
        wks['AN52'].border=thin_border
        wks['AO52'].border=thin_border
        wks['AP52'].border=thin_border
        wks['AQ52'].border=thin_border

        wks.merge_cells('AR52:AX52')
        wks['AR52']=wagetax_dec.revenu_ni2
        wks['AR52'].border=thin_border
        wks['AS52'].border=thin_border
        wks['AT52'].border=thin_border
        wks['AU52'].border=thin_border
        wks['AV52'].border=thin_border
        wks['AW52'].border=thin_border
        wks['AX52'].border=thin_border

        wks.merge_cells('B57:C57')
        wks['B57'].border=thin_border
        wks['C57'].border=thin_border

        wks.merge_cells('D57:O57')
        wks['D57'].border=thin_border
        wks['E57'].border=thin_border
        wks['F57'].border=thin_border
        wks['G57'].border=thin_border
        wks['H57'].border=thin_border
        wks['I57'].border=thin_border
        wks['J57'].border=thin_border
        wks['K57'].border=thin_border
        wks['L57'].border=thin_border
        wks['M57'].border=thin_border
        wks['N57'].border=thin_border
        wks['O57'].border=thin_border

        wks.merge_cells('P57:T57')
        wks['P57'].border=thin_border
        wks['Q57'].border=thin_border
        wks['R57'].border=thin_border
        wks['S57'].border=thin_border
        wks['T57'].border=thin_border

        wks.merge_cells('U57:V57')
        wks['U57'].border=thin_border
        wks['V57'].border=thin_border

        wks.merge_cells('W57:AF57')
        wks['W57'].border=thin_border
        wks['X57'].border=thin_border
        wks['Y57'].border=thin_border
        wks['Z57'].border=thin_border
        wks['AA57'].border=thin_border
        wks['AB57'].border=thin_border
        wks['AC57'].border=thin_border
        wks['AD57'].border=thin_border
        wks['AE57'].border=thin_border
        wks['AF57'].border=thin_border

        wks.merge_cells('AG57:AK57')
        wks['AG57'].border=thin_border
        wks['AH57'].border=thin_border
        wks['AI57'].border=thin_border
        wks['AJ57'].border=thin_border
        wks['AK57'].border=thin_border

        wks.merge_cells('AL57:AQ57')
        wks['AL57'].border=thin_border
        wks['AM57'].border=thin_border
        wks['AN57'].border=thin_border
        wks['AO57'].border=thin_border
        wks['AP57'].border=thin_border
        wks['AQ57'].border=thin_border

        wks.merge_cells('AR57:AX57')
        wks['AR57'].border=thin_border
        wks['AS57'].border=thin_border
        wks['AT57'].border=thin_border
        wks['AU57'].border=thin_border
        wks['AV57'].border=thin_border
        wks['AW57'].border=thin_border
        wks['AX57'].border=thin_border



        wks.merge_cells('B58:C58')
        wks['B58'].border=thin_border
        wks['C58'].border=thin_border

        wks.merge_cells('D58:O58')
        wks['D58'].border=thin_border
        wks['E58'].border=thin_border
        wks['F58'].border=thin_border
        wks['G58'].border=thin_border
        wks['H58'].border=thin_border
        wks['I58'].border=thin_border
        wks['J58'].border=thin_border
        wks['K58'].border=thin_border
        wks['L58'].border=thin_border
        wks['M58'].border=thin_border
        wks['N58'].border=thin_border
        wks['O58'].border=thin_border

        wks.merge_cells('P58:T58')
        wks['P58']=wagetax_dec.amount_tv3
        wks['P58'].border=thin_border
        wks['Q58'].border=thin_border
        wks['R58'].border=thin_border
        wks['S58'].border=thin_border
        wks['T58'].border=thin_border

        wks.merge_cells('U58:V58')
        wks['U58']=wagetax_dec.workforce_3
        wks['U58'].border=thin_border
        wks['V58'].border=thin_border

        wks.merge_cells('W58:AF58')
        wks['W58']=wagetax_dec.amount_exo3
        wks['W58'].border=thin_border
        wks['X58'].border=thin_border
        wks['Y58'].border=thin_border
        wks['Z58'].border=thin_border
        wks['AA58'].border=thin_border
        wks['AB58'].border=thin_border
        wks['AC58'].border=thin_border
        wks['AD58'].border=thin_border
        wks['AE58'].border=thin_border
        wks['AF58'].border=thin_border

        wks.merge_cells('AG58:AK58')
        wks['AG58']=wagetax_dec.amount_tax3
        wks['AG58'].border=thin_border
        wks['AH58'].border=thin_border
        wks['AI58'].border=thin_border
        wks['AJ58'].border=thin_border
        wks['AK58'].border=thin_border

        wks.merge_cells('AL58:AQ58')
        wks['AL58']=wagetax_dec.taxreduction_3
        wks['AL58'].border=thin_border
        wks['AM58'].border=thin_border
        wks['AN58'].border=thin_border
        wks['AO58'].border=thin_border
        wks['AP58'].border=thin_border
        wks['AQ58'].border=thin_border

        wks.merge_cells('AR58:AX58')
        wks['AR58']=wagetax_dec.revenu_ni3
        wks['AR58'].border=thin_border
        wks['AS58'].border=thin_border
        wks['AT58'].border=thin_border
        wks['AU58'].border=thin_border
        wks['AV58'].border=thin_border
        wks['AW58'].border=thin_border
        wks['AX58'].border=thin_border


        wks.merge_cells('B59:C59')
        wks['B59'].border=thin_border
        wks['C59'].border=thin_border

        wks.merge_cells('D59:O59')
        wks['D59'].border=thin_border
        wks['E59'].border=thin_border
        wks['F59'].border=thin_border
        wks['G59'].border=thin_border
        wks['H59'].border=thin_border
        wks['I59'].border=thin_border
        wks['J59'].border=thin_border
        wks['K59'].border=thin_border
        wks['L59'].border=thin_border
        wks['M59'].border=thin_border
        wks['N59'].border=thin_border
        wks['O59'].border=thin_border

        wks.merge_cells('P59:T59')
        wks['P59']=wagetax_dec.amount_tv4
        wks['P59'].border=thin_border
        wks['Q59'].border=thin_border
        wks['R59'].border=thin_border
        wks['S59'].border=thin_border
        wks['T59'].border=thin_border

        wks.merge_cells('U59:V59')
        wks['U59']=wagetax_dec.workforce_4
        wks['U59'].border=thin_border
        wks['V59'].border=thin_border

        wks.merge_cells('W59:AF59')
        wks['W59']=wagetax_dec.amount_exo4
        wks['W59'].border=thin_border
        wks['X59'].border=thin_border
        wks['Y59'].border=thin_border
        wks['Z59'].border=thin_border
        wks['AA59'].border=thin_border
        wks['AB59'].border=thin_border
        wks['AC59'].border=thin_border
        wks['AD59'].border=thin_border
        wks['AE59'].border=thin_border
        wks['AF59'].border=thin_border

        wks.merge_cells('AG59:AK59')
        wks['AG59']=wagetax_dec.amount_tax4
        wks['AG59'].border=thin_border
        wks['AH59'].border=thin_border
        wks['AI59'].border=thin_border
        wks['AJ59'].border=thin_border
        wks['AK59'].border=thin_border

        wks.merge_cells('AL59:AQ59')
        wks['AL59']=wagetax_dec.taxreduction_4
        wks['AL59'].border=thin_border
        wks['AM59'].border=thin_border
        wks['AN59'].border=thin_border
        wks['AO59'].border=thin_border
        wks['AP59'].border=thin_border
        wks['AQ59'].border=thin_border

        wks.merge_cells('AR59:AX59')
        wks['AR59']=wagetax_dec.revenu_ni4
        wks['AR59'].border=thin_border
        wks['AS59'].border=thin_border
        wks['AT59'].border=thin_border
        wks['AU59'].border=thin_border
        wks['AV59'].border=thin_border
        wks['AW59'].border=thin_border
        wks['AX59'].border=thin_border


        wks.merge_cells('B64:C64')
        wks['B64'].border=thin_border
        wks['C64'].border=thin_border

        wks.merge_cells('D64:T64')
        wks['D64'].border=thin_border
        wks['E64'].border=thin_border
        wks['F64'].border=thin_border
        wks['G64'].border=thin_border
        wks['H64'].border=thin_border
        wks['I64'].border=thin_border
        wks['J64'].border=thin_border
        wks['K64'].border=thin_border
        wks['L64'].border=thin_border
        wks['M64'].border=thin_border
        wks['N64'].border=thin_border
        wks['O64'].border=thin_border
        wks['P64'].border=thin_border
        wks['Q64'].border=thin_border
        wks['R64'].border=thin_border
        wks['S64'].border=thin_border
        wks['T64'].border=thin_border

        wks.merge_cells('U64:AF64')
        wks['U64'].border=thin_border
        wks['V64'].border=thin_border
        wks['W64'].border=thin_border
        wks['X64'].border=thin_border
        wks['Y64'].border=thin_border
        wks['Z64'].border=thin_border
        wks['AA64'].border=thin_border
        wks['AB64'].border=thin_border
        wks['AC64'].border=thin_border
        wks['AD64'].border=thin_border
        wks['AE64'].border=thin_border
        wks['AF64'].border=thin_border

        wks.merge_cells('B65:C65')
        wks['B65'].border=thin_border
        wks['C65'].border=thin_border

        wks.merge_cells('D65:T65')
        wks['D65'].border=thin_border
        wks['E65'].border=thin_border
        wks['F65'].border=thin_border
        wks['G65'].border=thin_border
        wks['H65'].border=thin_border
        wks['I65'].border=thin_border
        wks['J65'].border=thin_border
        wks['K65'].border=thin_border
        wks['L65'].border=thin_border
        wks['M65'].border=thin_border
        wks['N65'].border=thin_border
        wks['O65'].border=thin_border
        wks['P65'].border=thin_border
        wks['Q65'].border=thin_border
        wks['R65'].border=thin_border
        wks['S65'].border=thin_border
        wks['T65'].border=thin_border

        wks.merge_cells('U65:AF65')
        wks['U65']=wagetax_dec.amount_rv
        wks['U65'].border=thin_border
        wks['V65'].border=thin_border
        wks['W65'].border=thin_border
        wks['X65'].border=thin_border
        wks['Y65'].border=thin_border
        wks['Z65'].border=thin_border
        wks['AA65'].border=thin_border
        wks['AB65'].border=thin_border
        wks['AC65'].border=thin_border
        wks['AD65'].border=thin_border
        wks['AE65'].border=thin_border
        wks['AF65'].border=thin_border

        wks.merge_cells('B66:C66')
        wks['B66'].border=thin_border
        wks['C66'].border=thin_border

        wks.merge_cells('D66:T66')
        wks['D66'].border=thin_border
        wks['E66'].border=thin_border
        wks['F66'].border=thin_border
        wks['G66'].border=thin_border
        wks['H66'].border=thin_border
        wks['I66'].border=thin_border
        wks['J66'].border=thin_border
        wks['K66'].border=thin_border
        wks['L66'].border=thin_border
        wks['M66'].border=thin_border
        wks['N66'].border=thin_border
        wks['O66'].border=thin_border
        wks['P66'].border=thin_border
        wks['Q66'].border=thin_border
        wks['R66'].border=thin_border
        wks['S66'].border=thin_border
        wks['T66'].border=thin_border

        wks.merge_cells('U66:AF66')
        wks['U66']=wagetax_dec.amount_an
        wks['U66'].border=thin_border
        wks['V66'].border=thin_border
        wks['W66'].border=thin_border
        wks['X66'].border=thin_border
        wks['Y66'].border=thin_border
        wks['Z66'].border=thin_border
        wks['AA66'].border=thin_border
        wks['AB66'].border=thin_border
        wks['AC66'].border=thin_border
        wks['AD66'].border=thin_border
        wks['AE66'].border=thin_border
        wks['AF66'].border=thin_border

        wks.merge_cells('B67:C67')
        wks['B67'].border=thin_border
        wks['C67'].border=thin_border

        wks.merge_cells('D67:T67')
        wks['D67'].border=thin_border
        wks['E67'].border=thin_border
        wks['F67'].border=thin_border
        wks['G67'].border=thin_border
        wks['H67'].border=thin_border
        wks['I67'].border=thin_border
        wks['J67'].border=thin_border
        wks['K67'].border=thin_border
        wks['L67'].border=thin_border
        wks['M67'].border=thin_border
        wks['N67'].border=thin_border
        wks['O67'].border=thin_border
        wks['P67'].border=thin_border
        wks['Q67'].border=thin_border
        wks['R67'].border=thin_border
        wks['S67'].border=thin_border
        wks['T67'].border=thin_border

        wks.merge_cells('U67:AF67')
        wks['U67']=wagetax_dec.amount_a
        wks['U67'].border=thin_border
        wks['V67'].border=thin_border
        wks['W67'].border=thin_border
        wks['X67'].border=thin_border
        wks['Y67'].border=thin_border
        wks['Z67'].border=thin_border
        wks['AA67'].border=thin_border
        wks['AB67'].border=thin_border
        wks['AC67'].border=thin_border
        wks['AD67'].border=thin_border
        wks['AE67'].border=thin_border
        wks['AF67'].border=thin_border



        wks.merge_cells('B68:C68')
        wks['B68'].border=thin_border
        wks['C68'].border=thin_border

        wks.merge_cells('D68:T68')
        wks['D68'].border=thin_border
        wks['E68'].border=thin_border
        wks['F68'].border=thin_border
        wks['G68'].border=thin_border
        wks['H68'].border=thin_border
        wks['I68'].border=thin_border
        wks['J68'].border=thin_border
        wks['K68'].border=thin_border
        wks['L68'].border=thin_border
        wks['M68'].border=thin_border
        wks['N68'].border=thin_border
        wks['O68'].border=thin_border
        wks['P68'].border=thin_border
        wks['Q68'].border=thin_border
        wks['R68'].border=thin_border
        wks['S68'].border=thin_border
        wks['T68'].border=thin_border

        wks.merge_cells('U68:AF68')
        wks['U68']=wagetax_dec.amount_tb
        wks['U68'].border=thin_border
        wks['V68'].border=thin_border
        wks['W68'].border=thin_border
        wks['X68'].border=thin_border
        wks['Y68'].border=thin_border
        wks['Z68'].border=thin_border
        wks['AA68'].border=thin_border
        wks['AB68'].border=thin_border
        wks['AC68'].border=thin_border
        wks['AD68'].border=thin_border
        wks['AE68'].border=thin_border
        wks['AF68'].border=thin_border

        wks.merge_cells('B70:C70')
        wks['B70'].border=thin_border
        wks['C70'].border=thin_border

        wks.merge_cells('D70:T70')
        wks['D70'].border=thin_border
        wks['E70'].border=thin_border
        wks['F70'].border=thin_border
        wks['G70'].border=thin_border
        wks['H70'].border=thin_border
        wks['I70'].border=thin_border
        wks['J70'].border=thin_border
        wks['K70'].border=thin_border
        wks['L70'].border=thin_border
        wks['M70'].border=thin_border
        wks['N70'].border=thin_border
        wks['O70'].border=thin_border
        wks['P70'].border=thin_border
        wks['Q70'].border=thin_border
        wks['R70'].border=thin_border
        wks['S70'].border=thin_border
        wks['T70'].border=thin_border

        wks.merge_cells('U70:AD70')
        wks['U70'].border=thin_border
        wks['V70'].border=thin_border
        wks['W70'].border=thin_border
        wks['X70'].border=thin_border
        wks['Y70'].border=thin_border
        wks['Z70'].border=thin_border

        wks.merge_cells('AA70:AF70')
        wks['AA70'].border=thin_border
        wks['AB70'].border=thin_border
        wks['AC70'].border=thin_border
        wks['AD70'].border=thin_border
        wks['AE70'].border=thin_border
        wks['AF70'].border=thin_border

        wks.merge_cells('B71:C71')
        wks['B71'].border=thin_border
        wks['C71'].border=thin_border

        wks.merge_cells('D71:T71')
        wks['D71'].border=thin_border
        wks['E71'].border=thin_border
        wks['F71'].border=thin_border
        wks['G71'].border=thin_border
        wks['H71'].border=thin_border
        wks['I71'].border=thin_border
        wks['J71'].border=thin_border
        wks['K71'].border=thin_border
        wks['L71'].border=thin_border
        wks['M71'].border=thin_border
        wks['N71'].border=thin_border
        wks['O71'].border=thin_border
        wks['P71'].border=thin_border
        wks['Q71'].border=thin_border
        wks['R71'].border=thin_border
        wks['S71'].border=thin_border
        wks['T71'].border=thin_border

        wks.merge_cells('U71:Z71')
        wks['U71']=wagetax_dec.wageamount_1
        wks['U71'].border=thin_border
        wks['V71'].border=thin_border
        wks['W71'].border=thin_border
        wks['X71'].border=thin_border
        wks['Y71'].border=thin_border
        wks['Z71'].border=thin_border

        wks.merge_cells('AA71:AF71')
        wks['AA71']=wagetax_dec.allowance_1
        wks['AA71'].border=thin_border
        wks['AB71'].border=thin_border
        wks['AC71'].border=thin_border
        wks['AD71'].border=thin_border
        wks['AE71'].border=thin_border
        wks['AF71'].border=thin_border


        wks.merge_cells('B72:C72')
        wks['B72'].border=thin_border
        wks['C72'].border=thin_border

        wks.merge_cells('D72:T72')
        wks['D72'].border=thin_border
        wks['E72'].border=thin_border
        wks['F72'].border=thin_border
        wks['G72'].border=thin_border
        wks['H72'].border=thin_border
        wks['I72'].border=thin_border
        wks['J72'].border=thin_border
        wks['K72'].border=thin_border
        wks['L72'].border=thin_border
        wks['M72'].border=thin_border
        wks['N72'].border=thin_border
        wks['O72'].border=thin_border
        wks['P72'].border=thin_border
        wks['Q72'].border=thin_border
        wks['R72'].border=thin_border
        wks['S72'].border=thin_border
        wks['T72'].border=thin_border

        wks.merge_cells('U72:Z72')
        wks['U72']=wagetax_dec.totalrevenu_1
        wks['U72'].border=thin_border
        wks['V72'].border=thin_border
        wks['W72'].border=thin_border
        wks['X72'].border=thin_border
        wks['Y72'].border=thin_border
        wks['Z72'].border=thin_border

        wks.merge_cells('AA72:AF72')
        wks['AA72']=wagetax_dec.netrevenu_1
        wks['AA72'].border=thin_border
        wks['AB72'].border=thin_border
        wks['AC72'].border=thin_border
        wks['AD72'].border=thin_border
        wks['AE72'].border=thin_border
        wks['AF72'].border=thin_border


        wks.merge_cells('B73:C73')
        wks['B73'].border=thin_border
        wks['C73'].border=thin_border

        wks.merge_cells('D73:T73')
        wks['D73'].border=thin_border
        wks['E73'].border=thin_border
        wks['F73'].border=thin_border
        wks['G73'].border=thin_border
        wks['H73'].border=thin_border
        wks['I73'].border=thin_border
        wks['J73'].border=thin_border
        wks['K73'].border=thin_border
        wks['L73'].border=thin_border
        wks['M73'].border=thin_border
        wks['N73'].border=thin_border
        wks['O73'].border=thin_border
        wks['P73'].border=thin_border
        wks['Q73'].border=thin_border
        wks['R73'].border=thin_border
        wks['S73'].border=thin_border
        wks['T73'].border=thin_border

        wks.merge_cells('U73:Z73')
        wks['U73']=wagetax_dec.wageamount_2
        wks['U73'].border=thin_border
        wks['V73'].border=thin_border
        wks['W73'].border=thin_border
        wks['X73'].border=thin_border
        wks['Y73'].border=thin_border
        wks['Z73'].border=thin_border

        wks.merge_cells('AA73:AF73')
        wks['AA73']=wagetax_dec.allowance_2
        wks['AA73'].border=thin_border
        wks['AB73'].border=thin_border
        wks['AC73'].border=thin_border
        wks['AD73'].border=thin_border
        wks['AE73'].border=thin_border
        wks['AF73'].border=thin_border


        wks.merge_cells('B74:C74')
        wks['B74'].border=thin_border
        wks['C74'].border=thin_border

        wks.merge_cells('D74:T74')
        wks['D74'].border=thin_border
        wks['E74'].border=thin_border
        wks['F74'].border=thin_border
        wks['G74'].border=thin_border
        wks['H74'].border=thin_border
        wks['I74'].border=thin_border
        wks['J74'].border=thin_border
        wks['K74'].border=thin_border
        wks['L74'].border=thin_border
        wks['M74'].border=thin_border
        wks['N74'].border=thin_border
        wks['O74'].border=thin_border
        wks['P74'].border=thin_border
        wks['Q74'].border=thin_border
        wks['R74'].border=thin_border
        wks['S74'].border=thin_border
        wks['T74'].border=thin_border

        wks.merge_cells('U74:Z74')
        wks['U74']=wagetax_dec.totalrevenu_2
        wks['U74'].border=thin_border
        wks['V74'].border=thin_border
        wks['W74'].border=thin_border
        wks['X74'].border=thin_border
        wks['Y74'].border=thin_border
        wks['Z74'].border=thin_border

        wks.merge_cells('AA74:AF74')
        wks['AA74']=wagetax_dec.netrevenu_2
        wks['AA74'].border=thin_border
        wks['AB74'].border=thin_border
        wks['AC74'].border=thin_border
        wks['AD74'].border=thin_border
        wks['AE74'].border=thin_border
        wks['AF74'].border=thin_border

        wks.merge_cells('B76:C76')
        wks['B76'].border=thin_border
        wks['C76'].border=thin_border

        wks.merge_cells('D76:T76')
        wks['D76'].border=thin_border
        wks['E76'].border=thin_border
        wks['F76'].border=thin_border
        wks['G76'].border=thin_border
        wks['H76'].border=thin_border
        wks['I76'].border=thin_border
        wks['J76'].border=thin_border
        wks['K76'].border=thin_border
        wks['L76'].border=thin_border
        wks['M76'].border=thin_border
        wks['N76'].border=thin_border
        wks['O76'].border=thin_border
        wks['P76'].border=thin_border
        wks['Q76'].border=thin_border
        wks['R76'].border=thin_border
        wks['S76'].border=thin_border
        wks['T76'].border=thin_border

        wks.merge_cells('U76:Z76')
        wks['U76']=wagetax_dec.totalnetamount_1
        wks['U76'].border=thin_border
        wks['V76'].border=thin_border
        wks['W76'].border=thin_border
        wks['X76'].border=thin_border
        wks['Y76'].border=thin_border
        wks['Z76'].border=thin_border

        wks.merge_cells('AA76:AF76')
        wks['AA76']=wagetax_dec.totalnetamount_2
        wks['AA76'].border=thin_border
        wks['AB76'].border=thin_border
        wks['AC76'].border=thin_border
        wks['AD76'].border=thin_border
        wks['AE76'].border=thin_border
        wks['AF76'].border=thin_border

        wks.merge_cells(start_row=89, start_column=2, end_row=90, end_column=3)
        wks['B89'].border=thin_border
        wks['C89'].border=thin_border
        wks['B90'].border=thin_border
        wks['C90'].border=thin_border

        wks.merge_cells(start_row=89, start_column=4, end_row=90, end_column=20)
        wks['D89'].border=thin_border
        wks['E89'].border=thin_border
        wks['F89'].border=thin_border
        wks['G89'].border=thin_border
        wks['H89'].border=thin_border
        wks['I89'].border=thin_border
        wks['J89'].border=thin_border
        wks['K89'].border=thin_border
        wks['L89'].border=thin_border
        wks['M89'].border=thin_border
        wks['N89'].border=thin_border
        wks['O89'].border=thin_border
        wks['P89'].border=thin_border
        wks['Q89'].border=thin_border
        wks['R89'].border=thin_border
        wks['S89'].border=thin_border
        wks['T89'].border=thin_border
        wks['D90'].border=thin_border
        wks['E90'].border=thin_border
        wks['F90'].border=thin_border
        wks['G90'].border=thin_border
        wks['H90'].border=thin_border
        wks['I90'].border=thin_border
        wks['J90'].border=thin_border
        wks['K90'].border=thin_border
        wks['L90'].border=thin_border
        wks['M90'].border=thin_border
        wks['N90'].border=thin_border
        wks['O90'].border=thin_border
        wks['P90'].border=thin_border
        wks['Q90'].border=thin_border
        wks['R90'].border=thin_border
        wks['S90'].border=thin_border
        wks['T90'].border=thin_border

        wks.merge_cells(start_row=89, start_column=21, end_row=90, end_column=30)
        wks['U89'].border=thin_border
        wks['V89'].border=thin_border
        wks['W89'].border=thin_border
        wks['X89'].border=thin_border
        wks['Y89'].border=thin_border
        wks['Z89'].border=thin_border
        wks['AA89'].border=thin_border
        wks['AB89'].border=thin_border
        wks['AC89'].border=thin_border
        wks['AD89'].border=thin_border
        wks['U90'].border=thin_border
        wks['V90'].border=thin_border
        wks['W90'].border=thin_border
        wks['X90'].border=thin_border
        wks['Y90'].border=thin_border
        wks['Z90'].border=thin_border
        wks['AA90'].border=thin_border
        wks['AB90'].border=thin_border
        wks['AC90'].border=thin_border
        wks['AD90'].border=thin_border

        wks.merge_cells(start_row=89, start_column=31, end_row=90, end_column=40)
        wks['AE89'].border=thin_border
        wks['AF89'].border=thin_border
        wks['AG89'].border=thin_border
        wks['AH89'].border=thin_border
        wks['AI89'].border=thin_border
        wks['AJ89'].border=thin_border
        wks['AK89'].border=thin_border
        wks['AL89'].border=thin_border
        wks['AM89'].border=thin_border
        wks['AN89'].border=thin_border
        wks['AE90'].border=thin_border
        wks['AF90'].border=thin_border
        wks['AG90'].border=thin_border
        wks['AH90'].border=thin_border
        wks['AI90'].border=thin_border
        wks['AJ90'].border=thin_border
        wks['AK90'].border=thin_border
        wks['AL90'].border=thin_border
        wks['AM90'].border=thin_border
        wks['AN90'].border=thin_border

        wks.merge_cells(start_row=89, start_column=41, end_row=90, end_column=50)
        wks['AO89'].border=thin_border
        wks['AP89'].border=thin_border
        wks['AQ89'].border=thin_border
        wks['AR89'].border=thin_border
        wks['AS89'].border=thin_border
        wks['AT89'].border=thin_border
        wks['AU89'].border=thin_border
        wks['AV89'].border=thin_border
        wks['AW89'].border=thin_border
        wks['AX89'].border=thin_border
        wks['AO90'].border=thin_border
        wks['AP90'].border=thin_border
        wks['AQ90'].border=thin_border
        wks['AR90'].border=thin_border
        wks['AS90'].border=thin_border
        wks['AT90'].border=thin_border
        wks['AU90'].border=thin_border
        wks['AV90'].border=thin_border
        wks['AW90'].border=thin_border
        wks['AX90'].border=thin_border

        wks.merge_cells('B91:C91')
        wks['B91'].border=thin_border
        wks['C91'].border=thin_border

        wks.merge_cells('D91:T91')
        wks['D91'].border=thin_border
        wks['E91'].border=thin_border
        wks['F91'].border=thin_border
        wks['G91'].border=thin_border
        wks['H91'].border=thin_border
        wks['I91'].border=thin_border
        wks['J91'].border=thin_border
        wks['K91'].border=thin_border
        wks['L91'].border=thin_border
        wks['M91'].border=thin_border
        wks['N91'].border=thin_border
        wks['O91'].border=thin_border
        wks['P91'].border=thin_border
        wks['Q91'].border=thin_border
        wks['R91'].border=thin_border
        wks['S91'].border=thin_border
        wks['T91'].border=thin_border

        wks.merge_cells('U91:AD91')
        wks['U91']=wagetax_dec.base_1
        wks['U91'].border=thin_border
        wks['V91'].border=thin_border
        wks['W91'].border=thin_border
        wks['X91'].border=thin_border
        wks['Y91'].border=thin_border
        wks['Z91'].border=thin_border
        wks['AA91'].border=thin_border
        wks['AB91'].border=thin_border
        wks['AC91'].border=thin_border
        wks['AD91'].border=thin_border

        wks.merge_cells('AE91:AN91')
        wks['AE91']=wagetax_dec.taxrate_1
        wks['AE91'].border=thin_border
        wks['AF91'].border=thin_border
        wks['AG91'].border=thin_border
        wks['AH91'].border=thin_border
        wks['AI91'].border=thin_border
        wks['AJ91'].border=thin_border
        wks['AK91'].border=thin_border
        wks['AL91'].border=thin_border
        wks['AM91'].border=thin_border
        wks['AN91'].border=thin_border

        wks.merge_cells('AO91:AX91')
        wks['AO91']=wagetax_dec.taxamount_1
        wks['AO91'].border=thin_border
        wks['AP91'].border=thin_border
        wks['AQ91'].border=thin_border
        wks['AR91'].border=thin_border
        wks['AS91'].border=thin_border
        wks['AT91'].border=thin_border
        wks['AU91'].border=thin_border
        wks['AV91'].border=thin_border
        wks['AW91'].border=thin_border
        wks['AX91'].border=thin_border


        wks.merge_cells('B92:C92')
        wks['B92'].border=thin_border
        wks['C92'].border=thin_border

        wks.merge_cells('D92:T92')
        wks['D92'].border=thin_border
        wks['E92'].border=thin_border
        wks['F92'].border=thin_border
        wks['G92'].border=thin_border
        wks['H92'].border=thin_border
        wks['I92'].border=thin_border
        wks['J92'].border=thin_border
        wks['K92'].border=thin_border
        wks['L92'].border=thin_border
        wks['M92'].border=thin_border
        wks['N92'].border=thin_border
        wks['O92'].border=thin_border
        wks['P92'].border=thin_border
        wks['Q92'].border=thin_border
        wks['R92'].border=thin_border
        wks['S92'].border=thin_border
        wks['T92'].border=thin_border

        wks.merge_cells('U92:AD92')
        wks['U92']=wagetax_dec.base_2
        wks['U92'].border=thin_border
        wks['V92'].border=thin_border
        wks['W92'].border=thin_border
        wks['X92'].border=thin_border
        wks['Y92'].border=thin_border
        wks['Z92'].border=thin_border
        wks['AA92'].border=thin_border
        wks['AB92'].border=thin_border
        wks['AC92'].border=thin_border
        wks['AD92'].border=thin_border

        wks.merge_cells('AE92:AN92')
        wks['AE92']=wagetax_dec.taxrate_2
        wks['AE92'].border=thin_border
        wks['AF92'].border=thin_border
        wks['AG92'].border=thin_border
        wks['AH92'].border=thin_border
        wks['AI92'].border=thin_border
        wks['AJ92'].border=thin_border
        wks['AK92'].border=thin_border
        wks['AL92'].border=thin_border
        wks['AM92'].border=thin_border
        wks['AN92'].border=thin_border

        wks.merge_cells('AO92:AX92')
        wks['AO92']=wagetax_dec.taxamount_2
        wks['AO92'].border=thin_border
        wks['AP92'].border=thin_border
        wks['AQ92'].border=thin_border
        wks['AR92'].border=thin_border
        wks['AS92'].border=thin_border
        wks['AT92'].border=thin_border
        wks['AU92'].border=thin_border
        wks['AV92'].border=thin_border
        wks['AW92'].border=thin_border
        wks['AX92'].border=thin_border

        wks.merge_cells('B93:C93')
        wks['B93'].border=thin_border
        wks['C93'].border=thin_border

        wks.merge_cells('D93:T93')
        wks['D93'].border=thin_border
        wks['E93'].border=thin_border
        wks['F93'].border=thin_border
        wks['G93'].border=thin_border
        wks['H93'].border=thin_border
        wks['I93'].border=thin_border
        wks['J93'].border=thin_border
        wks['K93'].border=thin_border
        wks['L93'].border=thin_border
        wks['M93'].border=thin_border
        wks['N93'].border=thin_border
        wks['O93'].border=thin_border
        wks['P93'].border=thin_border
        wks['Q93'].border=thin_border
        wks['R93'].border=thin_border
        wks['S93'].border=thin_border
        wks['T93'].border=thin_border

        wks.merge_cells('U93:AD93')
        wks['U93']=wagetax_dec.base_3
        wks['U93'].border=thin_border
        wks['V93'].border=thin_border
        wks['W93'].border=thin_border
        wks['X93'].border=thin_border
        wks['Y93'].border=thin_border
        wks['Z93'].border=thin_border
        wks['AA93'].border=thin_border
        wks['AB93'].border=thin_border
        wks['AC93'].border=thin_border
        wks['AD93'].border=thin_border

        wks.merge_cells('AE93:AN93')
        wks['AE93']=wagetax_dec.taxrate_3
        wks['AE93'].border=thin_border
        wks['AF93'].border=thin_border
        wks['AG93'].border=thin_border
        wks['AH93'].border=thin_border
        wks['AI93'].border=thin_border
        wks['AJ93'].border=thin_border
        wks['AK93'].border=thin_border
        wks['AL93'].border=thin_border
        wks['AM93'].border=thin_border
        wks['AN93'].border=thin_border

        wks.merge_cells('AO93:AX93')
        wks['AO93']=wagetax_dec.taxamount_3
        wks['AO93'].border=thin_border
        wks['AP93'].border=thin_border
        wks['AQ93'].border=thin_border
        wks['AR93'].border=thin_border
        wks['AS93'].border=thin_border
        wks['AT93'].border=thin_border
        wks['AU93'].border=thin_border
        wks['AV93'].border=thin_border
        wks['AW93'].border=thin_border
        wks['AX93'].border=thin_border

        wks.merge_cells('B94:C94')
        wks['B94'].border=thin_border
        wks['C94'].border=thin_border

        wks.merge_cells('D94:AN94')
        wks['D94'].border=thin_border
        wks['E94'].border=thin_border
        wks['F94'].border=thin_border
        wks['G94'].border=thin_border
        wks['H94'].border=thin_border
        wks['I94'].border=thin_border
        wks['J94'].border=thin_border
        wks['K94'].border=thin_border
        wks['L94'].border=thin_border
        wks['M94'].border=thin_border
        wks['N94'].border=thin_border
        wks['O94'].border=thin_border
        wks['P94'].border=thin_border
        wks['Q94'].border=thin_border
        wks['R94'].border=thin_border
        wks['S94'].border=thin_border
        wks['T94'].border=thin_border
        wks['U94'].border=thin_border
        wks['V94'].border=thin_border
        wks['W94'].border=thin_border
        wks['X94'].border=thin_border
        wks['Y94'].border=thin_border
        wks['Z94'].border=thin_border
        wks['AA94'].border=thin_border
        wks['AB94'].border=thin_border
        wks['AC94'].border=thin_border
        wks['AD94'].border=thin_border
        wks['AE94'].border=thin_border
        wks['AF94'].border=thin_border
        wks['AG94'].border=thin_border
        wks['AH94'].border=thin_border
        wks['AI94'].border=thin_border
        wks['AJ94'].border=thin_border
        wks['AK94'].border=thin_border
        wks['AL94'].border=thin_border
        wks['AM94'].border=thin_border
        wks['AN94'].border=thin_border

        wks.merge_cells('AO94:AX94')
        wks['AO94']=wagetax_dec.total_wr
        wks['AO94'].border=thin_border
        wks['AP94'].border=thin_border
        wks['AQ94'].border=thin_border
        wks['AR94'].border=thin_border
        wks['AS94'].border=thin_border
        wks['AT94'].border=thin_border
        wks['AU94'].border=thin_border
        wks['AV94'].border=thin_border
        wks['AW94'].border=thin_border
        wks['AX94'].border=thin_border

        wks.merge_cells(start_row=101, start_column=2, end_row=102, end_column=3)
        wks['B101'].border=thin_border
        wks['C101'].border=thin_border
        wks['B102'].border=thin_border
        wks['C102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=4, end_row=102, end_column=8)
        wks['D101'].border=thin_border
        wks['E101'].border=thin_border
        wks['F101'].border=thin_border
        wks['G101'].border=thin_border
        wks['H101'].border=thin_border
        wks['D102'].border=thin_border
        wks['E102'].border=thin_border
        wks['F102'].border=thin_border
        wks['G102'].border=thin_border
        wks['H102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=9, end_row=102, end_column=22)
        wks['I101'].border=thin_border
        wks['I102'].border=thin_border
        wks['J101'].border=thin_border
        wks['K101'].border=thin_border
        wks['L101'].border=thin_border
        wks['M101'].border=thin_border
        wks['N101'].border=thin_border
        wks['O101'].border=thin_border
        wks['J102'].border=thin_border
        wks['K102'].border=thin_border
        wks['L102'].border=thin_border
        wks['M102'].border=thin_border
        wks['N102'].border=thin_border
        wks['O102'].border=thin_border
        wks['P101'].border=thin_border
        wks['Q101'].border=thin_border
        wks['R101'].border=thin_border
        wks['S101'].border=thin_border
        wks['T101'].border=thin_border
        wks['U101'].border=thin_border
        wks['P102'].border=thin_border
        wks['Q102'].border=thin_border
        wks['R102'].border=thin_border
        wks['S102'].border=thin_border
        wks['T102'].border=thin_border
        wks['U102'].border=thin_border
        wks['V101'].border=thin_border
        wks['V102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=23, end_row=102, end_column=28)
        wks['W101'].border=thin_border
        wks['W102'].border=thin_border
        wks['X101'].border=thin_border
        wks['Y101'].border=thin_border
        wks['Z101'].border=thin_border
        wks['AA101'].border=thin_border
        wks['AB101'].border=thin_border
        wks['X102'].border=thin_border
        wks['Y102'].border=thin_border
        wks['Z102'].border=thin_border
        wks['AA102'].border=thin_border
        wks['AB102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=29, end_row=102, end_column=34)
        wks['AC101'].border=thin_border
        wks['AD101'].border=thin_border
        wks['AE101'].border=thin_border
        wks['AF101'].border=thin_border
        wks['AG101'].border=thin_border
        wks['AH101'].border=thin_border
        wks['AC102'].border=thin_border
        wks['AD102'].border=thin_border
        wks['AE102'].border=thin_border
        wks['AF102'].border=thin_border
        wks['AG102'].border=thin_border
        wks['AH102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=35, end_row=102, end_column=40)
        wks['AI101'].border=thin_border
        wks['AJ101'].border=thin_border
        wks['AK101'].border=thin_border
        wks['AL101'].border=thin_border
        wks['AM101'].border=thin_border
        wks['AI102'].border=thin_border
        wks['AJ102'].border=thin_border
        wks['AK102'].border=thin_border
        wks['AL102'].border=thin_border
        wks['AM102'].border=thin_border
        wks['AN102'].border=thin_border

        wks.merge_cells(start_row=101, start_column=41, end_row=102, end_column=54)
        wks['AO101'].border=thin_border
        wks['AP101'].border=thin_border
        wks['AQ101'].border=thin_border
        wks['AR101'].border=thin_border
        wks['AS101'].border=thin_border
        wks['AT101'].border=thin_border
        wks['AU101'].border=thin_border
        wks['AV101'].border=thin_border
        wks['AW101'].border=thin_border
        wks['AX101'].border=thin_border
        wks['AO102'].border=thin_border
        wks['AP102'].border=thin_border
        wks['AQ102'].border=thin_border
        wks['AR102'].border=thin_border
        wks['AS102'].border=thin_border
        wks['AT102'].border=thin_border
        wks['AU102'].border=thin_border
        wks['AV102'].border=thin_border
        wks['AW102'].border=thin_border
        wks['AX102'].border=thin_border

        wks.merge_cells('B103:C103')
        wks['B103'].border=thin_border
        wks['C103'].border=thin_border

        wks.merge_cells(start_row=103, start_column=4, end_row=105, end_column=8)
        wks['D103'].border=thin_border
        wks['E103'].border=thin_border
        wks['F103'].border=thin_border
        wks['G103'].border=thin_border
        wks['H103'].border=thin_border
        wks['D104'].border=thin_border
        wks['E104'].border=thin_border
        wks['F104'].border=thin_border
        wks['G104'].border=thin_border
        wks['H104'].border=thin_border
        wks['D105'].border=thin_border
        wks['E105'].border=thin_border
        wks['F105'].border=thin_border
        wks['G105'].border=thin_border
        wks['H105'].border=thin_border

        wks.merge_cells(start_row=106, start_column=4, end_row=107, end_column=8)
        wks['D106'].border=thin_border
        wks['E106'].border=thin_border
        wks['F106'].border=thin_border
        wks['G106'].border=thin_border
        wks['H106'].border=thin_border
        wks['D107'].border=thin_border
        wks['E107'].border=thin_border
        wks['F107'].border=thin_border
        wks['G107'].border=thin_border
        wks['H107'].border=thin_border

        wks.merge_cells(start_row=106, start_column=35, end_row=107, end_column=40)
        wks['AI106']=wagetax_dec.rate_e4
        wks['AI106'].border=thin_border
        wks['AJ106'].border=thin_border
        wks['AK106'].border=thin_border
        wks['AL106'].border=thin_border
        wks['AM106'].border=thin_border
        wks['AN106'].border=thin_border
        wks['AI107'].border=thin_border
        wks['AJ107'].border=thin_border
        wks['AK107'].border=thin_border
        wks['AL107'].border=thin_border
        wks['AM107'].border=thin_border
        wks['AN107'].border=thin_border

        wks.merge_cells('I103:V103')
        wks['I103'].border=thin_border
        wks['J103'].border=thin_border
        wks['K103'].border=thin_border
        wks['L103'].border=thin_border
        wks['M103'].border=thin_border
        wks['N103'].border=thin_border
        wks['O103'].border=thin_border
        wks['P103'].border=thin_border
        wks['Q103'].border=thin_border
        wks['R103'].border=thin_border
        wks['S103'].border=thin_border
        wks['T103'].border=thin_border
        wks['U103'].border=thin_border
        wks['V103'].border=thin_border

        wks.merge_cells('W103:AB103')
        wks['W103']=wagetax_dec.wageforce_e1
        wks['W103'].border=thin_border
        wks['X103'].border=thin_border
        wks['Y103'].border=thin_border
        wks['Z103'].border=thin_border
        wks['AA103'].border=thin_border
        wks['AB103'].border=thin_border

        wks.merge_cells('AC103:AH103')
        wks['AC103']=wagetax_dec.revenu_nii1
        wks['AC103'].border=thin_border
        wks['AD103'].border=thin_border
        wks['AE103'].border=thin_border
        wks['AF103'].border=thin_border
        wks['AG103'].border=thin_border
        wks['AH103'].border=thin_border

        wks.merge_cells('AI103:AN103')
        wks['AI103']=wagetax_dec.rate_e1
        wks['AI103'].border=thin_border
        wks['AJ103'].border=thin_border
        wks['AK103'].border=thin_border
        wks['AL103'].border=thin_border
        wks['AM103'].border=thin_border
        wks['AN103'].border=thin_border

        wks.merge_cells('AO103:AX103')
        wks['AO103']=wagetax_dec.amount_e1
        wks['AO103'].border=thin_border
        wks['AP103'].border=thin_border
        wks['AQ103'].border=thin_border
        wks['AR103'].border=thin_border
        wks['AS103'].border=thin_border
        wks['AT103'].border=thin_border
        wks['AU103'].border=thin_border
        wks['AV103'].border=thin_border
        wks['AW103'].border=thin_border
        wks['AX103'].border=thin_border

        wks.merge_cells('B104:C104')
        wks['B104'].border=thin_border
        wks['C104'].border=thin_border

        wks.merge_cells('I104:V104')
        wks['I104'].border=thin_border
        wks['J104'].border=thin_border
        wks['K104'].border=thin_border
        wks['L104'].border=thin_border
        wks['M104'].border=thin_border
        wks['N104'].border=thin_border
        wks['O104'].border=thin_border
        wks['P104'].border=thin_border
        wks['Q104'].border=thin_border
        wks['R104'].border=thin_border
        wks['S104'].border=thin_border
        wks['T104'].border=thin_border
        wks['U104'].border=thin_border

        wks.merge_cells('W104:AB104')
        wks['W104']=wagetax_dec.wageforce_e2
        wks['W104'].border=thin_border
        wks['X104'].border=thin_border
        wks['Y104'].border=thin_border
        wks['Z104'].border=thin_border
        wks['AA104'].border=thin_border
        wks['AB104'].border=thin_border

        wks.merge_cells('AC104:AH104')
        wks['AC104']=wagetax_dec.revenu_nii2
        wks['AC104'].border=thin_border
        wks['AD104'].border=thin_border
        wks['AE104'].border=thin_border
        wks['AF104'].border=thin_border
        wks['AG104'].border=thin_border
        wks['AH104'].border=thin_border

        wks.merge_cells('AI104:AN104')
        wks['AI104']=wagetax_dec.rate_e2
        wks['AI104'].border=thin_border
        wks['AJ104'].border=thin_border
        wks['AK104'].border=thin_border
        wks['AL104'].border=thin_border
        wks['AM104'].border=thin_border
        wks['AN104'].border=thin_border

        wks.merge_cells('AO104:AX104')
        wks['AO104']=wagetax_dec.amount_e2
        wks['AO104'].border=thin_border
        wks['AP104'].border=thin_border
        wks['AQ104'].border=thin_border
        wks['AR104'].border=thin_border
        wks['AS104'].border=thin_border
        wks['AT104'].border=thin_border
        wks['AU104'].border=thin_border
        wks['AV104'].border=thin_border
        wks['AW104'].border=thin_border
        wks['AX104'].border=thin_border

        wks.merge_cells('B105:C105')
        wks['B105'].border=thin_border
        wks['C105'].border=thin_border

        wks.merge_cells('I105:V105')
        wks['I105'].border=thin_border
        wks['J105'].border=thin_border
        wks['K105'].border=thin_border
        wks['L105'].border=thin_border
        wks['M105'].border=thin_border
        wks['N105'].border=thin_border
        wks['O105'].border=thin_border
        wks['P105'].border=thin_border
        wks['Q105'].border=thin_border
        wks['R105'].border=thin_border
        wks['S105'].border=thin_border
        wks['T105'].border=thin_border
        wks['U105'].border=thin_border
        wks['V105'].border=thin_border

        wks.merge_cells('W105:AB105')
        wks['W105']=wagetax_dec.workforce_e3
        wks['W105'].border=thin_border
        wks['X105'].border=thin_border
        wks['Y105'].border=thin_border
        wks['Z105'].border=thin_border
        wks['AA105'].border=thin_border
        wks['AB105'].border=thin_border

        wks.merge_cells('AC105:AH105')
        wks['AC105']=wagetax_dec.revenu_nii3
        wks['AC105'].border=thin_border
        wks['AD105'].border=thin_border
        wks['AE105'].border=thin_border
        wks['AF105'].border=thin_border
        wks['AG105'].border=thin_border
        wks['AH105'].border=thin_border

        wks.merge_cells('AI105:AN105')
        wks['AI105']=wagetax_dec.rate_e3
        wks['AI105'].border=thin_border
        wks['AJ105'].border=thin_border
        wks['AK105'].border=thin_border
        wks['AL105'].border=thin_border
        wks['AM105'].border=thin_border
        wks['AN105'].border=thin_border

        wks.merge_cells('AO105:AX105')
        wks['AO105']=wagetax_dec.amount_e3
        wks['AO105'].border=thin_border
        wks['AP105'].border=thin_border
        wks['AQ105'].border=thin_border
        wks['AR105'].border=thin_border
        wks['AS105'].border=thin_border
        wks['AT105'].border=thin_border
        wks['AU105'].border=thin_border
        wks['AV105'].border=thin_border
        wks['AW105'].border=thin_border
        wks['AX105'].border=thin_border

        wks.merge_cells('B106:C106')
        wks['B106'].border=thin_border
        wks['C106'].border=thin_border

        wks.merge_cells('I106:V106')
        wks['I106'].border=thin_border
        wks['J106'].border=thin_border
        wks['K106'].border=thin_border
        wks['L106'].border=thin_border
        wks['M106'].border=thin_border
        wks['N106'].border=thin_border
        wks['O106'].border=thin_border
        wks['P106'].border=thin_border
        wks['Q106'].border=thin_border
        wks['R106'].border=thin_border
        wks['S106'].border=thin_border
        wks['T106'].border=thin_border
        wks['U106'].border=thin_border
        wks['V106'].border=thin_border

        wks.merge_cells('W106:AB106')
        wks['U106']=wagetax_dec.workforce_e4
        wks['W106'].border=thin_border
        wks['X106'].border=thin_border
        wks['Y106'].border=thin_border
        wks['Z106'].border=thin_border
        wks['AA106'].border=thin_border
        wks['AB106'].border=thin_border

        wks.merge_cells('AC106:AH106')
        wks['AC106']=wagetax_dec.revenu_nii4
        wks['AC106'].border=thin_border
        wks['AD106'].border=thin_border
        wks['AE106'].border=thin_border
        wks['AF106'].border=thin_border
        wks['AG106'].border=thin_border
        wks['AH106'].border=thin_border

        wks.merge_cells('AO106:AX106')
        wks['AO106']=wagetax_dec.amount_e4
        wks['AO106'].border=thin_border
        wks['AP106'].border=thin_border
        wks['AQ106'].border=thin_border
        wks['AR106'].border=thin_border
        wks['AS106'].border=thin_border
        wks['AT106'].border=thin_border
        wks['AU106'].border=thin_border
        wks['AV106'].border=thin_border
        wks['AW106'].border=thin_border
        wks['AX106'].border=thin_border

        wks.merge_cells('B107:C107')
        wks['B107'].border=thin_border
        wks['C107'].border=thin_border

        wks.merge_cells('I107:V107')
        wks['I107'].border=thin_border
        wks['J107'].border=thin_border
        wks['K107'].border=thin_border
        wks['L107'].border=thin_border
        wks['M107'].border=thin_border
        wks['N107'].border=thin_border
        wks['O107'].border=thin_border
        wks['P107'].border=thin_border
        wks['Q107'].border=thin_border
        wks['R107'].border=thin_border
        wks['S107'].border=thin_border
        wks['T107'].border=thin_border
        wks['U107'].border=thin_border
        wks['V107'].border=thin_border

        wks.merge_cells('W107:AB107')
        wks['W107']=wagetax_dec.workforce_e5
        wks['W107'].border=thin_border
        wks['X107'].border=thin_border
        wks['Y107'].border=thin_border
        wks['Z107'].border=thin_border
        wks['AA107'].border=thin_border
        wks['AB107'].border=thin_border

        wks.merge_cells('AC107:AH107')
        wks['AC107']=wagetax_dec.revenu_nii5
        wks['AC107'].border=thin_border
        wks['AD107'].border=thin_border
        wks['AE107'].border=thin_border
        wks['AF107'].border=thin_border
        wks['AG107'].border=thin_border
        wks['AH107'].border=thin_border

        wks.merge_cells('AO107:AX107')
        wks['AO107']=wagetax_dec.amount_e5
        wks['AO107'].border=thin_border
        wks['AP107'].border=thin_border
        wks['AQ107'].border=thin_border
        wks['AR107'].border=thin_border
        wks['AS107'].border=thin_border
        wks['AT107'].border=thin_border
        wks['AU107'].border=thin_border
        wks['AV107'].border=thin_border
        wks['AW107'].border=thin_border
        wks['AX108'].border=thin_border

        wks.merge_cells('B108:C108')
        wks['B108'].border=thin_border
        wks['C108'].border=thin_border

        wks.merge_cells('D108:AN108')
        wks['D108'].border=thin_border
        wks['E108'].border=thin_border
        wks['F108'].border=thin_border
        wks['G108'].border=thin_border
        wks['H108'].border=thin_border
        wks['I108'].border=thin_border
        wks['J108'].border=thin_border
        wks['K108'].border=thin_border
        wks['L108'].border=thin_border
        wks['M108'].border=thin_border
        wks['N108'].border=thin_border
        wks['O108'].border=thin_border
        wks['P108'].border=thin_border
        wks['Q108'].border=thin_border
        wks['R108'].border=thin_border
        wks['S108'].border=thin_border
        wks['T108'].border=thin_border
        wks['U108'].border=thin_border
        wks['V108'].border=thin_border
        wks['W108'].border=thin_border
        wks['X108'].border=thin_border
        wks['Y108'].border=thin_border
        wks['Z108'].border=thin_border
        wks['AA108'].border=thin_border
        wks['AB108'].border=thin_border
        wks['AC108'].border=thin_border
        wks['AD108'].border=thin_border
        wks['AE108'].border=thin_border
        wks['AF108'].border=thin_border
        wks['AG108'].border=thin_border
        wks['AH108'].border=thin_border
        wks['AI108'].border=thin_border
        wks['AJ108'].border=thin_border
        wks['AK108'].border=thin_border
        wks['AL108'].border=thin_border
        wks['AM108'].border=thin_border
        wks['AN108'].border=thin_border

        wks.merge_cells('AO108:AX108')
        wks['AO108']=wagetax_dec.totalcontribution
        wks['AO108'].border=thin_border
        wks['AP108'].border=thin_border
        wks['AQ108'].border=thin_border
        wks['AR108'].border=thin_border
        wks['AS108'].border=thin_border
        wks['AT108'].border=thin_border
        wks['AU108'].border=thin_border
        wks['AV108'].border=thin_border
        wks['AW108'].border=thin_border
        wks['AX108'].border=thin_border

        wks.merge_cells(start_row=114, start_column=2, end_row=115, end_column=3)
        wks['B114'].border=thin_border
        wks['C114'].border=thin_border
        wks['B115'].border=thin_border
        wks['C115'].border=thin_border

        wks.merge_cells(start_row=114, start_column=4, end_row=115, end_column=30)
        wks['D114'].border=thin_border
        wks['E114'].border=thin_border
        wks['F114'].border=thin_border
        wks['G114'].border=thin_border
        wks['H114'].border=thin_border
        wks['I114'].border=thin_border
        wks['D115'].border=thin_border
        wks['E115'].border=thin_border
        wks['F115'].border=thin_border
        wks['G115'].border=thin_border
        wks['H115'].border=thin_border
        wks['I115'].border=thin_border
        wks['J114'].border=thin_border
        wks['K114'].border=thin_border
        wks['L114'].border=thin_border
        wks['M114'].border=thin_border
        wks['N114'].border=thin_border
        wks['O114'].border=thin_border
        wks['J115'].border=thin_border
        wks['K115'].border=thin_border
        wks['L115'].border=thin_border
        wks['M115'].border=thin_border
        wks['N115'].border=thin_border
        wks['O115'].border=thin_border
        wks['P114'].border=thin_border
        wks['Q114'].border=thin_border
        wks['R114'].border=thin_border
        wks['S114'].border=thin_border
        wks['T114'].border=thin_border
        wks['U114'].border=thin_border
        wks['P115'].border=thin_border
        wks['Q115'].border=thin_border
        wks['R115'].border=thin_border
        wks['S115'].border=thin_border
        wks['T115'].border=thin_border
        wks['U115'].border=thin_border
        wks['V114'].border=thin_border
        wks['W114'].border=thin_border
        wks['X114'].border=thin_border
        wks['Y114'].border=thin_border
        wks['Z114'].border=thin_border
        wks['AA114'].border=thin_border
        wks['AB114'].border=thin_border
        wks['V115'].border=thin_border
        wks['W115'].border=thin_border
        wks['X115'].border=thin_border
        wks['Y115'].border=thin_border
        wks['Z115'].border=thin_border
        wks['AA115'].border=thin_border
        wks['AB115'].border=thin_border

        wks.merge_cells(start_row=114, start_column=31, end_row=115, end_column=36)
        wks['AC114'].border=thin_border
        wks['AD114'].border=thin_border
        wks['AE114'].border=thin_border
        wks['AF114'].border=thin_border
        wks['AG114'].border=thin_border
        wks['AH114'].border=thin_border
        wks['AC115'].border=thin_border
        wks['AD115'].border=thin_border
        wks['AE115'].border=thin_border
        wks['AF115'].border=thin_border
        wks['AG115'].border=thin_border
        wks['AH115'].border=thin_border

        wks.merge_cells(start_row=114, start_column=37, end_row=115, end_column=42)
        wks['AI114'].border=thin_border
        wks['AJ114'].border=thin_border
        wks['AK114'].border=thin_border
        wks['AL114'].border=thin_border
        wks['AM114'].border=thin_border
        wks['AI115'].border=thin_border
        wks['AJ115'].border=thin_border
        wks['AK115'].border=thin_border
        wks['AL115'].border=thin_border
        wks['AM115'].border=thin_border
        wks['AN115'].border=thin_border

        wks.merge_cells(start_row=114, start_column=43, end_row=115, end_column=54)
        wks['AO114'].border=thin_border
        wks['AP114'].border=thin_border
        wks['AQ114'].border=thin_border
        wks['AR114'].border=thin_border
        wks['AS114'].border=thin_border
        wks['AT114'].border=thin_border
        wks['AU114'].border=thin_border
        wks['AV114'].border=thin_border
        wks['AW114'].border=thin_border
        wks['AX114'].border=thin_border
        wks['AN115'].border=thin_border
        wks['AO115'].border=thin_border
        wks['AP115'].border=thin_border
        wks['AQ115'].border=thin_border
        wks['AR115'].border=thin_border
        wks['AS115'].border=thin_border
        wks['AT115'].border=thin_border
        wks['AU115'].border=thin_border
        wks['AV115'].border=thin_border
        wks['AW115'].border=thin_border
        wks['AX115'].border=thin_border

        wks.merge_cells('B116:C116')
        wks['B116'].border=thin_border
        wks['C116'].border=thin_border

        wks.merge_cells('D116:AB116')
        wks['D116'].border=thin_border
        wks['E116'].border=thin_border
        wks['F116'].border=thin_border
        wks['G116'].border=thin_border
        wks['H116'].border=thin_border
        wks['I116'].border=thin_border
        wks['J116'].border=thin_border
        wks['K116'].border=thin_border
        wks['L116'].border=thin_border
        wks['M116'].border=thin_border
        wks['N116'].border=thin_border
        wks['O116'].border=thin_border
        wks['P116'].border=thin_border
        wks['Q116'].border=thin_border
        wks['R116'].border=thin_border
        wks['S116'].border=thin_border
        wks['T116'].border=thin_border
        wks['U116'].border=thin_border
        wks['V116'].border=thin_border
        wks['W116'].border=thin_border
        wks['X116'].border=thin_border
        wks['Y116'].border=thin_border
        wks['Z116'].border=thin_border
        wks['AA116'].border=thin_border
        wks['AB116'].border=thin_border

        wks.merge_cells('AC116:AH116')
        wks['AC116']=wagetax_dec.revenu_netimp
        wks['AC116'].border=thin_border
        wks['AD116'].border=thin_border
        wks['AE116'].border=thin_border
        wks['AF116'].border=thin_border
        wks['AG116'].border=thin_border
        wks['AH116'].border=thin_border

        wks.merge_cells('AI116:AN116')
        wks['AI116']=wagetax_dec.rate_re
        wks['AI116'].border=thin_border
        wks['AJ116'].border=thin_border
        wks['AK116'].border=thin_border
        wks['AL116'].border=thin_border
        wks['AM116'].border=thin_border
        wks['AN116'].border=thin_border

        wks.merge_cells('AO116:AX116')
        wks['AO116']=wagetax_dec.amount_retained
        wks['AO116'].border=thin_border
        wks['AP116'].border=thin_border
        wks['AQ116'].border=thin_border
        wks['AR116'].border=thin_border
        wks['AS116'].border=thin_border
        wks['AT116'].border=thin_border
        wks['AU116'].border=thin_border
        wks['AV116'].border=thin_border
        wks['AW116'].border=thin_border
        wks['AX116'].border=thin_border

        wks.merge_cells(start_row=120, start_column=2, end_row=121, end_column=3)
        wks['B120'].border=thin_border
        wks['C120'].border=thin_border
        wks['B121'].border=thin_border
        wks['C121'].border=thin_border

        wks.merge_cells(start_row=120, start_column=4, end_row=121, end_column=10)
        wks['D120'].border=thin_border
        wks['E120'].border=thin_border
        wks['F120'].border=thin_border
        wks['G120'].border=thin_border
        wks['H120'].border=thin_border
        wks['I120'].border=thin_border
        wks['D121'].border=thin_border
        wks['E121'].border=thin_border
        wks['F121'].border=thin_border
        wks['G121'].border=thin_border
        wks['H121'].border=thin_border
        wks['I121'].border=thin_border

        wks.merge_cells(start_row=120, start_column=11, end_row=121, end_column=17)
        wks['J120'].border=thin_border
        wks['K120'].border=thin_border
        wks['L120'].border=thin_border
        wks['M120'].border=thin_border
        wks['N120'].border=thin_border
        wks['O120'].border=thin_border
        wks['J121'].border=thin_border
        wks['K121'].border=thin_border
        wks['L121'].border=thin_border
        wks['M121'].border=thin_border
        wks['N121'].border=thin_border
        wks['O121'].border=thin_border


        wks.merge_cells(start_row=120, start_column=18, end_row=121, end_column=22)
        wks['P120'].border=thin_border
        wks['Q120'].border=thin_border
        wks['R120'].border=thin_border
        wks['S120'].border=thin_border
        wks['T120'].border=thin_border
        wks['U120'].border=thin_border
        wks['P121'].border=thin_border
        wks['Q121'].border=thin_border
        wks['R121'].border=thin_border
        wks['S121'].border=thin_border
        wks['T121'].border=thin_border
        wks['U121'].border=thin_border

        wks.merge_cells(start_row=120, start_column=23, end_row=121, end_column=30)
        wks['V120'].border=thin_border
        wks['W120'].border=thin_border
        wks['X120'].border=thin_border
        wks['Y120'].border=thin_border
        wks['Z120'].border=thin_border
        wks['AA120'].border=thin_border
        wks['AB120'].border=thin_border
        wks['V121'].border=thin_border
        wks['W121'].border=thin_border
        wks['X121'].border=thin_border
        wks['Y121'].border=thin_border
        wks['Z121'].border=thin_border
        wks['AA121'].border=thin_border
        wks['AB121'].border=thin_border

        wks.merge_cells(start_row=120, start_column=31, end_row=121, end_column=42)
        wks['AC120'].border=thin_border
        wks['AD120'].border=thin_border
        wks['AE120'].border=thin_border
        wks['AF120'].border=thin_border
        wks['AG120'].border=thin_border
        wks['AH120'].border=thin_border
        wks['AI120'].border=thin_border
        wks['AJ120'].border=thin_border
        wks['AK120'].border=thin_border
        wks['AL120'].border=thin_border
        wks['AM120'].border=thin_border
        wks['AC121'].border=thin_border
        wks['AD121'].border=thin_border
        wks['AE121'].border=thin_border
        wks['AF121'].border=thin_border
        wks['AG121'].border=thin_border
        wks['AH121'].border=thin_border
        wks['AI121'].border=thin_border
        wks['AJ121'].border=thin_border
        wks['AK121'].border=thin_border
        wks['AL121'].border=thin_border
        wks['AM121'].border=thin_border

        wks.merge_cells(start_row=120, start_column=43, end_row=121, end_column=54)
        wks['AN120'].border=thin_border
        wks['AO120'].border=thin_border
        wks['AP120'].border=thin_border
        wks['AQ120'].border=thin_border
        wks['AR120'].border=thin_border
        wks['AS120'].border=thin_border
        wks['AT120'].border=thin_border
        wks['AU120'].border=thin_border
        wks['AV120'].border=thin_border
        wks['AW120'].border=thin_border
        wks['AX120'].border=thin_border
        wks['AN121'].border=thin_border
        wks['AO121'].border=thin_border
        wks['AP121'].border=thin_border
        wks['AQ121'].border=thin_border
        wks['AR121'].border=thin_border
        wks['AS121'].border=thin_border
        wks['AT121'].border=thin_border
        wks['AU121'].border=thin_border
        wks['AV121'].border=thin_border
        wks['AW121'].border=thin_border
        wks['AX121'].border=thin_border


        wks.merge_cells('B122:C122')
        wks['B122'].border=thin_border
        wks['C122'].border=thin_border

        wks.merge_cells('D122:I122')
        wks['D122']=wagetax_dec.tax
        wks['D122'].border=thin_border
        wks['E122'].border=thin_border
        wks['F122'].border=thin_border
        wks['G122'].border=thin_border
        wks['H122'].border=thin_border
        wks['I122'].border=thin_border

        wks.merge_cells('J122:O122')
        wks['J122']=wagetax_dec.contribution_n
        wks['J122'].border=thin_border
        wks['K122'].border=thin_border
        wks['L122'].border=thin_border
        wks['M122'].border=thin_border
        wks['N122'].border=thin_border
        wks['O122'].border=thin_border

        wks.merge_cells('P122:U122')
        wks['P122']=wagetax_dec.tax_gsr
        wks['P122'].border=thin_border
        wks['Q122'].border=thin_border
        wks['R122'].border=thin_border
        wks['S122'].border=thin_border
        wks['T122'].border=thin_border
        wks['U122'].border=thin_border

        wks.merge_cells('V122:AB122')
        wks['V122']=wagetax_dec.contribution_e
        wks['V122'].border=thin_border
        wks['W122'].border=thin_border
        wks['X122'].border=thin_border
        wks['Y122'].border=thin_border
        wks['Z122'].border=thin_border
        wks['AA122'].border=thin_border
        wks['AB122'].border=thin_border

        wks.merge_cells('AC122:AN122')
        wks['AC122']=wagetax_dec.contribution_nce
        wks['AC122'].border=thin_border
        wks['AD122'].border=thin_border
        wks['AE122'].border=thin_border
        wks['AF122'].border=thin_border
        wks['AG122'].border=thin_border
        wks['AH122'].border=thin_border
        wks['AI122'].border=thin_border
        wks['AJ122'].border=thin_border
        wks['AK122'].border=thin_border
        wks['AL122'].border=thin_border
        wks['AM122'].border=thin_border
        wks['AN122'].border=thin_border

        wks.merge_cells('AO122:AX122')
        wks['AO122']=wagetax_dec.total_r
        wks['AO122'].border=thin_border
        wks['AP122'].border=thin_border
        wks['AQ122'].border=thin_border
        wks['AR122'].border=thin_border
        wks['AS122'].border=thin_border
        wks['AT122'].border=thin_border
        wks['AU122'].border=thin_border
        wks['AV122'].border=thin_border
        wks['AW122'].border=thin_border
        wks['AX122'].border=thin_border


        wks.merge_cells('B124:C124')
        wks['B124'].border=thin_border
        wks['C124'].border=thin_border

        wks.merge_cells('D124:AD124')
        wks['D124'].border=thin_border
        wks['E124'].border=thin_border
        wks['F124'].border=thin_border
        wks['G124'].border=thin_border
        wks['H124'].border=thin_border
        wks['I124'].border=thin_border
        wks['J124'].border=thin_border
        wks['K124'].border=thin_border
        wks['L124'].border=thin_border
        wks['M124'].border=thin_border
        wks['N124'].border=thin_border
        wks['O124'].border=thin_border
        wks['P124'].border=thin_border
        wks['Q124'].border=thin_border
        wks['R124'].border=thin_border
        wks['S124'].border=thin_border
        wks['T124'].border=thin_border
        wks['U124'].border=thin_border
        wks['V124'].border=thin_border
        wks['W124'].border=thin_border
        wks['X124'].border=thin_border
        wks['Y124'].border=thin_border
        wks['Z124'].border=thin_border
        wks['AA124'].border=thin_border
        wks['AB124'].border=thin_border
        wks['AC124'].border=thin_border
        wks['AD124'].border=thin_border

        wks.merge_cells('AE124:AX124')
        wks['AE124']=wagetax_dec.amount_tp
        wks['AE124'].border=thin_border
        wks['AF124'].border=thin_border
        wks['AG124'].border=thin_border
        wks['AH124'].border=thin_border
        wks['AI124'].border=thin_border
        wks['AJ124'].border=thin_border
        wks['AK124'].border=thin_border
        wks['AL124'].border=thin_border
        wks['AM124'].border=thin_border
        wks['AN124'].border=thin_border
        wks['AO124'].border=thin_border
        wks['AP124'].border=thin_border
        wks['AQ124'].border=thin_border
        wks['AR124'].border=thin_border
        wks['AS124'].border=thin_border
        wks['AT124'].border=thin_border
        wks['AU124'].border=thin_border
        wks['AV124'].border=thin_border
        wks['AW124'].border=thin_border
        wks['AX124'].border=thin_border

        
        wks.merge_cells('B125:C125')
        wks['B125'].border=thin_border
        wks['C125'].border=thin_border

        wks.merge_cells('D125:AD125')
        wks['D125'].border=thin_border
        wks['E125'].border=thin_border
        wks['F125'].border=thin_border
        wks['G125'].border=thin_border
        wks['H125'].border=thin_border
        wks['I125'].border=thin_border
        wks['J125'].border=thin_border
        wks['K125'].border=thin_border
        wks['L125'].border=thin_border
        wks['M125'].border=thin_border
        wks['N125'].border=thin_border
        wks['O125'].border=thin_border
        wks['P125'].border=thin_border
        wks['Q125'].border=thin_border
        wks['R125'].border=thin_border
        wks['S125'].border=thin_border
        wks['T125'].border=thin_border
        wks['U125'].border=thin_border
        wks['V125'].border=thin_border
        wks['W125'].border=thin_border
        wks['X125'].border=thin_border
        wks['Y125'].border=thin_border
        wks['Z125'].border=thin_border
        wks['AA125'].border=thin_border
        wks['AB125'].border=thin_border
        wks['AC125'].border=thin_border
        wks['AD125'].border=thin_border

        wks.merge_cells('AE125:AX125')
        wks['AE125']=wagetax_dec.amount_tr
        wks['AE125'].border=thin_border
        wks['AF125'].border=thin_border
        wks['AG125'].border=thin_border
        wks['AH125'].border=thin_border
        wks['AI125'].border=thin_border
        wks['AJ125'].border=thin_border
        wks['AK125'].border=thin_border
        wks['AL125'].border=thin_border
        wks['AM125'].border=thin_border
        wks['AN125'].border=thin_border
        wks['AO125'].border=thin_border
        wks['AP125'].border=thin_border
        wks['AQ125'].border=thin_border
        wks['AR125'].border=thin_border
        wks['AS125'].border=thin_border
        wks['AT125'].border=thin_border
        wks['AU125'].border=thin_border
        wks['AV125'].border=thin_border
        wks['AW125'].border=thin_border
        wks['AX125'].border=thin_border

       

        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'wagetax_report_xls','wagetax_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'wagetax.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }