# -*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *

class trainingfund_declaration(osv.osv):
    _name="trainingfund.declaration"
    _inherit="trainingfund.declaration"


    def trainingfund_report_xls(self, cr, uid, ids, context=None): # fonction report its

        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\fdfp_template.xlsx"
        trainingfund_dec=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path)
        wks = wbk.active

        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )

        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )

        dashed_border = Border(
            bottom=Side(style='dashed')
            )

        wks.merge_cells('A6:E6')
        wks['A6'].border=thin_border
        wks['B6'].border=thin_border
        wks['C6'].border=thin_border
        wks['D6'].border=thin_border
        wks['E6'].border=thin_border

        wks.merge_cells('V6:Z6')
        wks['V6'].border=thin_border
        wks['W6'].border=thin_border
        wks['X6'].border=thin_border
        wks['Y6'].border=thin_border
        wks['Z6'].border=thin_border

        wks.merge_cells('V17:AA17')
        wks['V17'].border=dashed_border
        wks['W17'].border=dashed_border
        wks['X17'].border=dashed_border
        wks['Y17'].border=dashed_border
        wks['Z17'].border=dashed_border
        wks['AA17'].border=dashed_border

        wks.merge_cells('J21:M21')
        wks['J21'].border=thin_border
        wks['K21'].border=thin_border
        wks['L21'].border=thin_border
        wks['M21'].border=thin_border

        wks.merge_cells('A25:K25')
        wks['A25'].border=thin_border
        wks['B25'].border=thin_border
        wks['C25'].border=thin_border
        wks['D25'].border=thin_border
        wks['E25'].border=thin_border
        wks['F25'].border=thin_border
        wks['G25'].border=thin_border
        wks['H25'].border=thin_border
        wks['I25'].border=thin_border
        wks['J25'].border=thin_border
        wks['K25'].border=thin_border

        wks.merge_cells('L25:S25')
        wks['L25'].border=thin_border
        wks['M25'].border=thin_border
        wks['N25'].border=thin_border
        wks['O25'].border=thin_border
        wks['P25'].border=thin_border
        wks['Q25'].border=thin_border
        wks['R25'].border=thin_border
        wks['S25'].border=thin_border

        wks.merge_cells('T25:U25')
        wks['T25'].border=thin_border
        wks['U25'].border=thin_border

        wks.merge_cells('V25:AA25')
        wks['V25'].border=thin_border
        wks['W25'].border=thin_border
        wks['X25'].border=thin_border
        wks['Y25'].border=thin_border
        wks['Z25'].border=thin_border
        wks['AA25'].border=thin_border

        wks.merge_cells('A26:K26')
        wks['A26'].border=thin_border
        wks['B26'].border=thin_border
        wks['C26'].border=thin_border
        wks['D26'].border=thin_border
        wks['E26'].border=thin_border
        wks['F26'].border=thin_border
        wks['G26'].border=thin_border
        wks['H26'].border=thin_border
        wks['I26'].border=thin_border
        wks['J26'].border=thin_border
        wks['K26'].border=thin_border

        wks.merge_cells('L26:S26')
        wks['L26'].border=thin_border
        wks['M26'].border=thin_border
        wks['N26'].border=thin_border
        wks['O26'].border=thin_border
        wks['P26'].border=thin_border
        wks['Q26'].border=thin_border
        wks['R26'].border=thin_border
        wks['S26'].border=thin_border

        wks.merge_cells('T26:U26')
        wks['T26'].border=thin_border
        wks['U26'].border=thin_border

        wks.merge_cells('V26:AA26')
        wks['V26'].border=thin_border
        wks['W26'].border=thin_border
        wks['X26'].border=thin_border
        wks['Y26'].border=thin_border
        wks['Z26'].border=thin_border
        wks['AA26'].border=thin_border

        wks.merge_cells('A27:K27')
        wks['A27'].border=thin_border
        wks['B27'].border=thin_border
        wks['C27'].border=thin_border
        wks['D27'].border=thin_border
        wks['E27'].border=thin_border
        wks['F27'].border=thin_border
        wks['G27'].border=thin_border
        wks['H27'].border=thin_border
        wks['I27'].border=thin_border
        wks['J27'].border=thin_border
        wks['K27'].border=thin_border

        wks.merge_cells('L27:S27')
        wks['L27'].border=thin_border
        wks['M27'].border=thin_border
        wks['N27'].border=thin_border
        wks['O27'].border=thin_border
        wks['P27'].border=thin_border
        wks['Q27'].border=thin_border
        wks['R27'].border=thin_border
        wks['S27'].border=thin_border

        wks.merge_cells('T27:U27')
        wks['T27'].border=thin_border
        wks['U27'].border=thin_border

        wks.merge_cells('V27:AA27')
        wks['V27'].border=thin_border
        wks['W27'].border=thin_border
        wks['X27'].border=thin_border
        wks['Y27'].border=thin_border
        wks['Z27'].border=thin_border
        wks['AA27'].border=thin_border

        wks.merge_cells(start_row=20, start_column=29, end_row=26, end_column=30)
        wks['T29'].border=thin_border
        wks['U29'].border=thin_border
        wks['V29'].border=thin_border
        wks['W29'].border=thin_border
        wks['X29'].border=thin_border
        wks['Y29'].border=thin_border
        wks['Z29'].border=thin_border
        wks['T30'].border=thin_border
        wks['U30'].border=thin_border
        wks['V30'].border=thin_border
        wks['W30'].border=thin_border
        wks['X30'].border=thin_border
        wks['Y30'].border=thin_border
        wks['Z30'].border=thin_border

        wks.merge_cells(start_row=1, start_column=35, end_row=17, end_column=36)
        wks['A35'].border=thin_border
        wks['B35'].border=thin_border
        wks['C35'].border=thin_border
        wks['D35'].border=thin_border
        wks['E35'].border=thin_border
        wks['F35'].border=thin_border
        wks['G35'].border=thin_border
        wks['H35'].border=thin_border
        wks['I35'].border=thin_border
        wks['J35'].border=thin_border
        wks['K35'].border=thin_border
        wks['L35'].border=thin_border
        wks['M35'].border=thin_border
        wks['N35'].border=thin_border
        wks['O35'].border=thin_border
        wks['P35'].border=thin_border
        wks['Q35'].border=thin_border
        wks['A36'].border=thin_border
        wks['B36'].border=thin_border
        wks['C36'].border=thin_border
        wks['D36'].border=thin_border
        wks['E36'].border=thin_border
        wks['F36'].border=thin_border
        wks['G36'].border=thin_border
        wks['H36'].border=thin_border
        wks['I36'].border=thin_border
        wks['J36'].border=thin_border
        wks['K36'].border=thin_border
        wks['L36'].border=thin_border
        wks['M36'].border=thin_border
        wks['N36'].border=thin_border
        wks['O36'].border=thin_border
        wks['P36'].border=thin_border
        wks['Q36'].border=thin_border

        wks.merge_cells('R35:Z35')
        wks['R35'].border=thin_border
        wks['S35'].border=thin_border
        wks['T35'].border=thin_border
        wks['U35'].border=thin_border
        wks['V35'].border=thin_border
        wks['W35'].border=thin_border
        wks['X35'].border=thin_border
        wks['Y35'].border=thin_border
        wks['Z35'].border=thin_border

        wks.merge_cells('R36:Z36')
        wks['R36'].border=thin_border
        wks['S36'].border=thin_border
        wks['T36'].border=thin_border
        wks['U36'].border=thin_border
        wks['V36'].border=thin_border
        wks['W36'].border=thin_border
        wks['X36'].border=thin_border
        wks['Y36'].border=thin_border
        wks['Z36'].border=thin_border


        wks.merge_cells('A37:Q37')
        wks['A37'].border=thin_border
        wks['B37'].border=thin_border
        wks['C37'].border=thin_border
        wks['D37'].border=thin_border
        wks['E37'].border=thin_border
        wks['F37'].border=thin_border
        wks['G37'].border=thin_border
        wks['H37'].border=thin_border
        wks['I37'].border=thin_border
        wks['J37'].border=thin_border
        wks['K37'].border=thin_border
        wks['L37'].border=thin_border
        wks['M37'].border=thin_border
        wks['N37'].border=thin_border
        wks['O37'].border=thin_border
        wks['P37'].border=thin_border
        wks['Q37'].border=thin_border


        wks.merge_cells('R37:Z37')
        wks['R37'].border=thin_border
        wks['S37'].border=thin_border
        wks['T37'].border=thin_border
        wks['U37'].border=thin_border
        wks['V37'].border=thin_border
        wks['W37'].border=thin_border
        wks['X37'].border=thin_border
        wks['Y37'].border=thin_border
        wks['Z37'].border=thin_border

        wks.merge_cells('A38:Q38')
        wks['A38'].border=thin_border
        wks['B38'].border=thin_border
        wks['C38'].border=thin_border
        wks['D38'].border=thin_border
        wks['E38'].border=thin_border
        wks['F38'].border=thin_border
        wks['G38'].border=thin_border
        wks['H38'].border=thin_border
        wks['I38'].border=thin_border
        wks['J38'].border=thin_border
        wks['K38'].border=thin_border
        wks['L38'].border=thin_border
        wks['M38'].border=thin_border
        wks['N38'].border=thin_border
        wks['O38'].border=thin_border
        wks['P38'].border=thin_border
        wks['Q38'].border=thin_border


        wks.merge_cells('R38:Z38')
        wks['R38'].border=thin_border
        wks['S38'].border=thin_border
        wks['T38'].border=thin_border
        wks['U38'].border=thin_border
        wks['V38'].border=thin_border
        wks['W38'].border=thin_border
        wks['X38'].border=thin_border
        wks['Y38'].border=thin_border
        wks['Z38'].border=thin_border

        wks.merge_cells('A39:Q39')
        wks['A39'].border=thin_border
        wks['B39'].border=thin_border
        wks['C39'].border=thin_border
        wks['D39'].border=thin_border
        wks['E39'].border=thin_border
        wks['F39'].border=thin_border
        wks['G39'].border=thin_border
        wks['H39'].border=thin_border
        wks['I39'].border=thin_border
        wks['J39'].border=thin_border
        wks['K39'].border=thin_border
        wks['L39'].border=thin_border
        wks['M39'].border=thin_border
        wks['N39'].border=thin_border
        wks['O39'].border=thin_border
        wks['P39'].border=thin_border
        wks['Q39'].border=thin_border


        wks.merge_cells('R39:Z39')
        wks['R39'].border=thin_border
        wks['S39'].border=thin_border
        wks['T39'].border=thin_border
        wks['U39'].border=thin_border
        wks['V39'].border=thin_border
        wks['W39'].border=thin_border
        wks['X39'].border=thin_border
        wks['Y39'].border=thin_border
        wks['Z39'].border=thin_border

        wks.merge_cells(start_row=1, start_column=40, end_row=17, end_column=41)
        wks['A40'].border=thin_border
        wks['B40'].border=thin_border
        wks['C40'].border=thin_border
        wks['D40'].border=thin_border
        wks['E40'].border=thin_border
        wks['F40'].border=thin_border
        wks['G40'].border=thin_border
        wks['H40'].border=thin_border
        wks['I40'].border=thin_border
        wks['J40'].border=thin_border
        wks['K40'].border=thin_border
        wks['L40'].border=thin_border
        wks['M40'].border=thin_border
        wks['N40'].border=thin_border
        wks['O40'].border=thin_border
        wks['P40'].border=thin_border
        wks['Q40'].border=thin_border
        wks['A41'].border=thin_border
        wks['B41'].border=thin_border
        wks['C41'].border=thin_border
        wks['D41'].border=thin_border
        wks['E41'].border=thin_border
        wks['F41'].border=thin_border
        wks['G41'].border=thin_border
        wks['H41'].border=thin_border
        wks['I41'].border=thin_border
        wks['J41'].border=thin_border
        wks['K41'].border=thin_border
        wks['L41'].border=thin_border
        wks['M41'].border=thin_border
        wks['N41'].border=thin_border
        wks['O41'].border=thin_border
        wks['P41'].border=thin_border
        wks['Q41'].border=thin_border

        wks.merge_cells(start_row=18, start_column=40, end_row=26, end_column=41)
        wks['R40'].border=thin_border
        wks['S40'].border=thin_border
        wks['T40'].border=thin_border
        wks['U40'].border=thin_border
        wks['V40'].border=thin_border
        wks['W40'].border=thin_border
        wks['X40'].border=thin_border
        wks['Y40'].border=thin_border
        wks['Z40'].border=thin_border
        wks['R41'].border=thin_border
        wks['S41'].border=thin_border
        wks['T41'].border=thin_border
        wks['U41'].border=thin_border
        wks['V41'].border=thin_border
        wks['W41'].border=thin_border
        wks['X41'].border=thin_border
        wks['Y41'].border=thin_border
        wks['Z41'].border=thin_border
        
        wks.merge_cells('R43:Z43')
        wks['R43'].border=medium_border
        wks['S43'].border=medium_border
        wks['T43'].border=medium_border
        wks['U43'].border=medium_border
        wks['V43'].border=medium_border
        wks['W43'].border=medium_border
        wks['X43'].border=medium_border
        wks['Y43'].border=medium_border
        wks['Z43'].border=medium_border

        wks.merge_cells('R45:Z45')
        wks['R45'].border=medium_border
        wks['S45'].border=medium_border
        wks['T45'].border=medium_border
        wks['U45'].border=medium_border
        wks['V45'].border=medium_border
        wks['W45'].border=medium_border
        wks['X45'].border=medium_border
        wks['Y45'].border=medium_border
        wks['Z45'].border=medium_border
        

        
        wks['D1']=trainingfund_dec.company_tax_code
        wks['V17']=trainingfund_dec.tax_service
        wks['J21']=trainingfund_dec.employe_workforce
        wks['L26']=trainingfund_dec.remuneration_bttaf
        wks['L27']=trainingfund_dec.remuneration_bttf
        wks['V17']=trainingfund_dec.tax_service
        wks['T26']=trainingfund_dec.rate_taf
        wks['T27']=trainingfund_dec.rate_tfcf
        wks['V26']=trainingfund_dec.monthly_amount_tatf
        wks['V27']=trainingfund_dec.regulation_state_sf
        wks['T29']=trainingfund_dec.total_amount_tfcsf
        wks['Q32']=trainingfund_dec.regulation_state_sf
        wks['R36']=trainingfund_dec.employe_workforce_asf
        wks['R37']=trainingfund_dec.amount_tcfsf
        wks['R38']=trainingfund_dec.amount_pay_tcfsf
        wks['R43']=trainingfund_dec.amount_tax_tcfsf
        wks['R39']=trainingfund_dec.engagement_tpsf
        wks['R41']=trainingfund_dec.payment
        wks['R45']=trainingfund_dec.payment_todo

        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'trainingfund_report_xls','trainingfund_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'trainingfund.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }