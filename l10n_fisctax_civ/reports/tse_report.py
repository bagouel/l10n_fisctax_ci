# -*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *


class eq_specialtax_declaration(osv.osv):
    _name="eq.specialtax.declaration"
    _inherit="eq.specialtax.declaration"

    def eq_specialtax_report_xls(self, cr, uid, ids, context=None): # fonction report tse

        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\tse_template.xlsx"
        eq_specialtax_dec=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path)
        wks = wbk.active

        thin1_border = Border(
            right=Side(style='thin'),  
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin2_border = Border(
            bottom=Side(style='thin')
            )
        thin8_border = Border(
            right=Side(style='thin'),
            bottom=Side(style='thin')
            )
        thin3_border = Border(
            left=Side(style='thin')
            )

        thin4_border = Border(
            bottom=Side(style='thin'),
            left=Side(style='thin')
            )
        thin5_border = Border(
            top=Side(style='thin'),
            right=Side(style='thin')
            )
        thin6_border = Border(
            right=Side(style='thin')
            )
        thin7_border = Border( 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        medium2_border = Border(
            right=Side(style='medium'), 
            )
        medium3_border = Border(
            right=Side(style='medium'),
            bottom=Side(style='thin') 
            )
        medium4_border = Border(
            left=Side(style='medium'),
            bottom=Side(style='thin') 
            )
        medium5_border = Border(
            left=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        medium6_border = Border(
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            )
        medium7_border = Border(
            left=Side(style='medium'),  
            top=Side(style='medium'), 
            )
        medium8_border = Border(
            right=Side(style='medium'),  
            bottom=Side(style='medium'), 
            )
        medium10_border = Border(
            bottom=Side(style='medium')
            )
        dashed_border = Border(
            bottom=Side(style='dashed')
            )
        melange_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='dashed')
            )
        melange2_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            )
        melange3_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),  
            bottom=Side(style='dashed')
            )

        wks.merge_cells('A6:E6')
        wks['A6'].border=thin_border
        wks['B6'].border=thin_border
        wks['C6'].border=thin_border
        wks['D6'].border=thin_border
        wks['E6'].border=thin_border

        wks.merge_cells('V6:Z6')
        wks['V6'].border=thin_border
        wks['W6'].border=thin_border
        wks['X6'].border=thin_border
        wks['Y6'].border=thin_border
        wks['Z6'].border=thin_border

        wks.merge_cells('V16:AA16')
        wks['V16'].border=dashed_border
        wks['W16'].border=dashed_border
        wks['X16'].border=dashed_border
        wks['Y16'].border=dashed_border
        wks['Z16'].border=dashed_border
        wks['AA16'].border=dashed_border

        wks.merge_cells('A21:J21')
        wks['A21'].border=thin_border
        wks['B21'].border=thin_border
        wks['C21'].border=thin_border
        wks['D21'].border=thin_border
        wks['E21'].border=thin_border
        wks['F21'].border=thin_border
        wks['G21'].border=thin_border
        wks['H21'].border=thin_border
        wks['I21'].border=thin_border
        wks['J21'].border=thin_border

        wks.merge_cells(start_row=22, start_column=1, end_row=24, end_column=10)
        wks['A22'].border=thin_border
        wks['B22'].border=thin_border
        wks['C22'].border=thin_border
        wks['D22'].border=thin_border
        wks['E22'].border=thin_border
        wks['F22'].border=thin_border
        wks['G22'].border=thin_border
        wks['H22'].border=thin_border
        wks['I22'].border=thin_border
        wks['J22'].border=thin_border
        wks['A23'].border=thin_border
        wks['B23'].border=thin_border
        wks['C23'].border=thin_border
        wks['D23'].border=thin_border
        wks['E23'].border=thin_border
        wks['F23'].border=thin_border
        wks['G23'].border=thin_border
        wks['H23'].border=thin_border
        wks['I23'].border=thin_border
        wks['J23'].border=thin_border
        wks['A24'].border=thin_border
        wks['B24'].border=thin_border
        wks['C24'].border=thin_border
        wks['D24'].border=thin_border
        wks['E24'].border=thin_border
        wks['F24'].border=thin_border
        wks['G24'].border=thin_border
        wks['H24'].border=thin_border
        wks['I24'].border=thin_border
        wks['J24'].border=thin_border

        wks.merge_cells('A26:J26')
        wks['A26'].border=thin_border
        wks['B26'].border=thin_border
        wks['C26'].border=thin_border
        wks['D26'].border=thin_border
        wks['E26'].border=thin_border
        wks['F26'].border=thin_border
        wks['G26'].border=thin_border
        wks['H26'].border=thin_border
        wks['I26'].border=thin_border
        wks['J26'].border=thin_border

        wks.merge_cells('A27:J27')
        wks['A27'].border=thin_border
        wks['B27'].border=thin_border
        wks['C27'].border=thin_border
        wks['D27'].border=thin_border
        wks['E27'].border=thin_border
        wks['F27'].border=thin_border
        wks['G27'].border=thin_border
        wks['H27'].border=thin_border
        wks['I27'].border=thin_border
        wks['J27'].border=thin_border

        wks.merge_cells('A28:J28')
        wks['A28'].border=thin_border
        wks['B28'].border=thin_border
        wks['C28'].border=thin_border
        wks['D28'].border=thin_border
        wks['E28'].border=thin_border
        wks['F28'].border=thin_border
        wks['G28'].border=thin_border
        wks['H28'].border=thin_border
        wks['I28'].border=thin_border
        wks['J28'].border=thin_border

        wks.merge_cells('A29:J29')
        wks['A29'].border=thin_border
        wks['B29'].border=thin_border
        wks['C29'].border=thin_border
        wks['D29'].border=thin_border
        wks['E29'].border=thin_border
        wks['F29'].border=thin_border
        wks['G29'].border=thin_border
        wks['H29'].border=thin_border
        wks['I29'].border=thin_border
        wks['J29'].border=thin_border

        wks.merge_cells('A30:J30')
        wks['A30'].border=thin_border
        wks['B30'].border=thin_border
        wks['C30'].border=thin_border
        wks['D30'].border=thin_border
        wks['E30'].border=thin_border
        wks['F30'].border=thin_border
        wks['G30'].border=thin_border
        wks['H30'].border=thin_border
        wks['I30'].border=thin_border
        wks['J30'].border=thin_border

        wks.merge_cells('V21:Z21')
        wks['V21'].border=thin_border
        wks['W21'].border=thin_border
        wks['X21'].border=thin_border
        wks['Y21'].border=thin_border
        wks['Z21'].border=thin_border

        wks.merge_cells(start_row=22, start_column=22, end_row=26, end_column=34)
        wks['V22'].border=thin_border
        wks['W22'].border=thin_border
        wks['X22'].border=thin_border
        wks['Y22'].border=thin_border
        wks['Z22'].border=thin_border
        wks['V23'].border=thin_border
        wks['W23'].border=thin_border
        wks['X23'].border=thin_border
        wks['Y23'].border=thin_border
        wks['Z23'].border=thin_border
        wks['V24'].border=thin_border
        wks['W24'].border=thin_border
        wks['X24'].border=thin_border
        wks['Y24'].border=thin_border
        wks['Z24'].border=thin_border
        wks['V25'].border=thin_border
        wks['W25'].border=thin_border
        wks['X25'].border=thin_border
        wks['Y25'].border=thin_border
        wks['Z25'].border=thin_border
        wks['V26'].border=thin_border
        wks['W26'].border=thin_border
        wks['X26'].border=thin_border
        wks['Y26'].border=thin_border
        wks['Z26'].border=thin_border
        wks['V27'].border=thin_border
        wks['W27'].border=thin_border
        wks['X27'].border=thin_border
        wks['Y27'].border=thin_border
        wks['Z27'].border=thin_border
        wks['V28'].border=thin_border
        wks['W28'].border=thin_border
        wks['X28'].border=thin_border
        wks['Y28'].border=thin_border
        wks['Z28'].border=thin_border
        wks['V29'].border=thin_border
        wks['W29'].border=thin_border
        wks['X29'].border=thin_border
        wks['Y29'].border=thin_border
        wks['Z29'].border=thin_border
        wks['V30'].border=thin_border
        wks['W30'].border=thin_border
        wks['X30'].border=thin_border
        wks['Y30'].border=thin_border
        wks['Z30'].border=thin_border
        wks['V31'].border=thin_border
        wks['W31'].border=thin_border
        wks['X31'].border=thin_border
        wks['Y31'].border=thin_border
        wks['Z31'].border=thin_border
        wks['V32'].border=thin_border
        wks['W32'].border=thin_border
        wks['X32'].border=thin_border
        wks['Y32'].border=thin_border
        wks['Z32'].border=thin_border
        wks['V33'].border=thin_border
        wks['W33'].border=thin_border
        wks['X33'].border=thin_border
        wks['Y33'].border=thin_border
        wks['Z33'].border=thin_border
        wks['V34'].border=thin_border
        wks['W34'].border=thin_border
        wks['X34'].border=thin_border
        wks['Y34'].border=thin_border
        wks['Z34'].border=thin_border

        wks.merge_cells('K21:T21')
        wks['K21'].border=melange_border
        wks['L21'].border=melange_border
        wks['M21'].border=melange_border
        wks['N21'].border=melange_border
        wks['O21'].border=melange_border
        wks['P21'].border=melange_border
        wks['Q21'].border=melange_border
        wks['R21'].border=melange_border
        wks['S21'].border=melange_border
        wks['T21'].border=melange_border

        wks.merge_cells('K22:T22')
        wks['K22'].border=melange2_border
        wks['L22'].border=melange2_border
        wks['M22'].border=melange2_border
        wks['N22'].border=melange2_border
        wks['O22'].border=melange2_border
        wks['P22'].border=melange2_border
        wks['Q22'].border=melange2_border
        wks['R22'].border=melange2_border
        wks['S22'].border=melange2_border
        wks['T22'].border=melange2_border

        wks.merge_cells('K23:T23')
        wks['K23'].border=melange3_border
        wks['L23'].border=melange3_border
        wks['M23'].border=melange3_border
        wks['N23'].border=melange3_border
        wks['O23'].border=melange3_border
        wks['P23'].border=melange3_border
        wks['Q23'].border=melange3_border
        wks['R23'].border=melange3_border
        wks['S23'].border=melange3_border
        wks['T23'].border=melange3_border

        wks.merge_cells('K24:T24')
        wks['K24'].border=melange3_border
        wks['L24'].border=melange3_border
        wks['M24'].border=melange3_border
        wks['N24'].border=melange3_border
        wks['O24'].border=melange3_border
        wks['P24'].border=melange3_border
        wks['Q24'].border=melange3_border
        wks['R24'].border=melange3_border
        wks['S24'].border=melange3_border
        wks['T24'].border=melange3_border

        wks.merge_cells('K26:T26')
        wks['K26'].border=thin_border
        wks['L26'].border=thin_border
        wks['M26'].border=thin_border
        wks['N26'].border=thin_border
        wks['O26'].border=thin_border
        wks['P26'].border=thin_border
        wks['Q26'].border=thin_border
        wks['R26'].border=thin_border
        wks['S26'].border=thin_border
        wks['T26'].border=thin_border

        wks.merge_cells('K27:T27')
        wks['K27'].border=thin_border
        wks['L27'].border=thin_border
        wks['M27'].border=thin_border
        wks['N27'].border=thin_border
        wks['O27'].border=thin_border
        wks['P27'].border=thin_border
        wks['Q27'].border=thin_border
        wks['R27'].border=thin_border
        wks['S27'].border=thin_border
        wks['T27'].border=thin_border

        wks.merge_cells('K28:T28')
        wks['K28'].border=thin_border
        wks['L28'].border=thin_border
        wks['M28'].border=thin_border
        wks['N28'].border=thin_border
        wks['O28'].border=thin_border
        wks['P28'].border=thin_border
        wks['Q28'].border=thin_border
        wks['R28'].border=thin_border
        wks['S28'].border=thin_border
        wks['T28'].border=thin_border

        wks.merge_cells('K29:T29')
        wks['K29'].border=thin_border
        wks['L29'].border=thin_border
        wks['M29'].border=thin_border
        wks['N29'].border=thin_border
        wks['O29'].border=thin_border
        wks['P29'].border=thin_border
        wks['Q29'].border=thin_border
        wks['R29'].border=thin_border
        wks['S29'].border=thin_border
        wks['T29'].border=thin_border

        wks.merge_cells('K30:T30')
        wks['K30'].border=thin_border
        wks['L30'].border=thin_border
        wks['M30'].border=thin_border
        wks['N30'].border=thin_border
        wks['O30'].border=thin_border
        wks['P30'].border=thin_border
        wks['Q30'].border=thin_border
        wks['R30'].border=thin_border
        wks['S30'].border=thin_border
        wks['T30'].border=thin_border

        wks.merge_cells('K32:T32')
        wks['K32'].border=medium_border
        wks['L32'].border=medium_border
        wks['M32'].border=medium_border
        wks['N32'].border=medium_border
        wks['O32'].border=medium_border
        wks['P32'].border=medium_border
        wks['Q32'].border=medium_border
        wks['R32'].border=medium_border
        wks['S32'].border=medium_border
        wks['T32'].border=medium_border

        wks.merge_cells('K34:T34')
        wks['K34'].border=medium_border
        wks['L34'].border=medium_border
        wks['M34'].border=medium_border
        wks['N34'].border=medium_border
        wks['O34'].border=medium_border
        wks['P34'].border=medium_border
        wks['Q34'].border=medium_border
        wks['R34'].border=medium_border
        wks['S34'].border=medium_border
        wks['T34'].border=medium_border

        
        wks['X9']=eq_specialtax_dec.date
        wks['D1']=eq_specialtax_dec.company_tax_code
        wks['V16']=eq_specialtax_dec.tax_service
        wks['K21']=eq_specialtax_dec.revenu_wout_tax
        wks['K23']=eq_specialtax_dec.ops_exempt_tax_revenu
        wks['K24']=eq_specialtax_dec.delivery_onself_revenu
        wks['K26']=eq_specialtax_dec.taxable_revenu
        wks['K27']=eq_specialtax_dec.rate
        wks['K28']=eq_specialtax_dec.tax_amount
        wks['K30']=eq_specialtax_dec.regulation
        wks['K32']=eq_specialtax_dec.amount_topay
        wks['K34']=eq_specialtax_dec.amount_toreport
        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'eq_specialtax_report_xls','eq_specialtax_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'eq_specialtax.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }