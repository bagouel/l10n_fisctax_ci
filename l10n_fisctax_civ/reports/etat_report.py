# -*- coding: utf-8 -*-
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *



class deductiblevat_statement(osv.osv):
    _name="deductiblevat.statement"
    _inherit="deductiblevat.statement"



    def deductiblevat_report_xls(self, cr, uid, ids, context=None): # fonction report etat

        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\etat_template.xlsx"
        deductiblevat_stat=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path)
        wks = wbk.active

        def insert_total_statement_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            lib={}
            tex={}
            font={}
            styl={}
            text1_sty={}
            text2={}
            text2_sty={}
            text3={}
            text3_sty={}

            for statement in self.browse(cr, uid, ids, context=context):
                val=6
                for line in statement.element_statement_ids:
                    val+=1
                    vals=val+1
                    valss=vals+1
                   

                col_start='I'+str(val)
                col_start1='A'+str(val)
                col_start3='A'+str(vals)
                col_start4='A'+str(valss)
                col_start2='H'+str(val)
                col_end='I'+str(vals)
                

                wks.merge_cells(start_row=val, start_column=9, end_row=vals, end_column=9)
                wks[col_start].border=thin_border
                wks[col_end].border=thin_border
                wks[col_start]=statement.total_statement
                wks[col_start].font=font2

                wks.merge_cells(start_row=val, start_column=8, end_row=vals, end_column=8)
                wks[col_start2]="TOTAL"
                wks[col_start2].font=font1

                wks[col_start1]="     (1) Pour les importations, indiquer le pays d’origine."
                wks[col_start1].font=font3

                wks[col_start3]="     (2) Indiquer la nature exacte et précise de chaque bien et prestation de service de façon individualisée."
                wks[col_start3].font=font3

                wks[col_start4]="      Sont à proscrire, les mentions d'ordre général telles que marchandises diverses, matériels de bureau, consommables…"
                wks[col_start4].font=font3

                res[statement.id]=wks[col_start].border  
                cont[statement.id]=wks[col_start]
                lib[statement.id]=wks[col_start2]

                font[statement.id]=wks[col_start2].font
                styl[statement.id]=wks[col_start].font

                tex[statement.id]=wks[col_start1]
                text1_sty[statement.id]=wks[col_start1].font
                
                text2[statement.id]=wks[col_start3]
                text2_sty[statement.id]=wks[col_start3].font

                text3[statement.id]=wks[col_start4]
                text3_sty[statement.id]=wks[col_start4].font

            return res , cont, lib, tex, font, styl, text2, text3, text1_sty, text2_sty, text3_sty


        def insert_date_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='A'+str(val)
                    wks[vals]= line.date
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont

        def insert_supplier_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='B'+str(val)
                    wks[vals]= line.supplier
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont

        def insert_supplier_taxt_code_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='C'+str(val)
                    wks[vals]= line.supplier_taxt_code
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont

        def insert_document_ref_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='D'+str(val)
                    wks[vals]= line.document_ref
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont    

        def insert_goods_definition_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='E'+str(val)
                    wks[vals]= line.goods_definition
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont    

        def insert_amount_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='F'+str(val)
                    wks[vals]= line.amount
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont    


        def insert_vat_amount_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='G'+str(val)
                    wks[vals]= line.vat_amount
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont    

        def insert_deduction_rate_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='H'+str(val)
                    wks[vals]= line.deduction_rate
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont  

        def insert_deductible_vat_amount_of_element_statement(self, cr, uid, ids, context=None):
            res={}
            cont={}
            for statement in self.browse(cr, uid, ids, context=context):
                val=5
                for line in statement.element_statement_ids:
                    val+=1
                    vals='I'+str(val)
                    wks[vals]= line.deductible_vat_amount
                    wks[vals].border=thin_border
                    res[statement.id]= wks[vals]
                    cont[statement.id]=wks[vals].border
            return res , cont 

        font = Font(name='Calibri',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,color='FF000000')
   

        thin1_border = Border(
            right=Side(style='thin'),  
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin2_border = Border(
            left=Side(style='thin')
            )
        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        dashed_border = Border(
            bottom=Side(style='dashed')
            )
        font1 = Font(
            name='Arial',
            size=14,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00000000')

        font2 = Font(
            name='Arial',
            size=13,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00000000')
        font3 = Font(
            name='Arial',
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00000000')
        
        wks['C1']=deductiblevat_stat.company_tax_code
        wks['F1']=deductiblevat_stat.tax_service
        wks['E2']=deductiblevat_stat.tax_period

        wks.merge_cells('B4:C4')
        wks['B4'].border=thin_border
        wks['C4'].border=thin_border
       
        wks.merge_cells(start_row=4, start_column=4, end_row=5, end_column=4)
        wks['D4'].border=thin_border
        wks['D5'].border=thin_border
        
        wks.merge_cells(start_row=4, start_column=5, end_row=5, end_column=5)
        wks['E4'].border=thin_border
        wks['E5'].border=thin_border

        wks.merge_cells(start_row=4, start_column=6, end_row=5, end_column=6)
        wks['F4'].border=thin_border
        wks['F5'].border=thin_border

        wks.merge_cells(start_row=4, start_column=7, end_row=5, end_column=7)
        wks['G4'].border=thin_border
        wks['G5'].border=thin_border

        wks.merge_cells(start_row=4, start_column=8, end_row=5, end_column=8)
        wks['H4'].border=thin_border
        wks['H5'].border=thin_border

        wks.merge_cells(start_row=4, start_column=9, end_row=5, end_column=9)
        wks['I4'].border=thin_border
        wks['I5'].border=thin_border

        insert_date_of_element_statement(self, cr, uid, ids, context=None)
        insert_supplier_of_element_statement(self, cr, uid, ids, context=None)
        insert_supplier_taxt_code_of_element_statement(self, cr, uid, ids, context=None)
        insert_document_ref_of_element_statement(self, cr, uid, ids, context=None)
        insert_goods_definition_of_element_statement(self, cr, uid, ids, context=None)
        insert_amount_of_element_statement(self, cr, uid, ids, context=None)
        insert_vat_amount_of_element_statement(self, cr, uid, ids, context=None)
        insert_deduction_rate_of_element_statement(self, cr, uid, ids, context=None)
        insert_deductible_vat_amount_of_element_statement(self, cr, uid, ids, context=None)

        insert_total_statement_of_element_statement(self, cr, uid, ids, context=None)
        
        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'deductiblevat_report_xls','deductiblevat_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'deductiblevat.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }
    



