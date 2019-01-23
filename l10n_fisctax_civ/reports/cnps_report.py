# -*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side 
from openerp.osv import  osv 
from openerp import models, fields, api, _
from datetime import datetime
from lxml import etree
from cStringIO import StringIO
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openpyxl, platform, os.path, xlwt, time, calendar, base64, logging, openerp.addons.decimal_precision as dp
from openpyxl import load_workbook
from openpyxl import workbook
from openerp.modules.module import *


class socialfund_declaration(osv.osv):
    _name="socialfund.declaration"
    _inherit="socialfund.declaration"


    def socialfund_report_xls(self, cr, uid, ids, context=None): # fonction report socialfund

    
        module_path=get_module_path('l10n_fisctax_civ')+"\\templates\\cnps_template.xlsx"
        socialfund_dec=self.browse(cr, uid, ids)
        fl = StringIO()
        if context is None :
                context={}
        wbk = openpyxl.load_workbook(module_path) 
        wks = wbk.active

        thin1_border = Border(
            right=Side(style='thin'),  
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        thin2_border = Border(
            left=Side(style='thin')
            )
        medium_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
            )
        dashed_border = Border(
            bottom=Side(style='dashed')
            )

        

        wks['O5']=socialfund_dec.company

        wks['O7']=socialfund_dec.address
        wks['AC7']=socialfund_dec.phone
        wks['S8']="TOTAL SALAIRES BRUTS PAYES AU COURS DE LA PERIODE" 

        wks.merge_cells('AG8:AM8')
        wks['AG8']=socialfund_dec.totalamount
        wks['AG8'].border=thin_border
        wks['AH8'].border=thin_border
        wks['AI8'].border=thin_border
        wks['AJ8'].border=thin_border
        wks['AK8'].border=thin_border
        wks['AL8'].border=thin_border
        wks['AM8'].border=thin_border

        wks.merge_cells('A7:B7')
        wks['A8']=socialfund_dec.code_ets
        wks.merge_cells('A8:B8')
        wks['A7'].border=thin_border
        wks['B7'].border=thin_border
        wks['A8'].border=thin_border
        wks['B8'].border=thin_border

        wks.merge_cells('C7:D7')
        wks['C8']=socialfund_dec.activity_code
        wks.merge_cells('C8:D8')
        wks['C7'].border=thin_border
        wks['D7'].border=thin_border
        wks['C8'].border=thin_border
        wks['D8'].border=thin_border

        wks.merge_cells('E7:G7')
        wks['E8']=socialfund_dec.employer
        wks.merge_cells('E8:G8')
        wks['E7'].border=thin_border
        wks['F7'].border=thin_border
        wks['G7'].border=thin_border
        wks['E8'].border=thin_border
        wks['F8'].border=thin_border
        wks['G8'].border=thin_border

        wks.merge_cells('H7:I7')
        wks['H8']=socialfund_dec.tax_period
        wks.merge_cells('H8:I8')
        wks['H7'].border=thin_border
        wks['I7'].border=thin_border
        wks['H8'].border=thin_border
        wks['I8'].border=thin_border

        wks.merge_cells('M10:S10')
        wks['M10']=socialfund_dec.month
        wks['N10'].border=thin1_border
        wks['O10'].border=thin1_border
        wks['P10'].border=thin1_border
        wks['Q10'].border=thin1_border
        wks['R10'].border=thin1_border
        wks['S10'].border=thin1_border

        wks.merge_cells('W10:AC10')
        wks['W10']=socialfund_dec.month2
        wks['W10'].border=thin1_border
        wks['X10'].border=thin1_border
        wks['Y10'].border=thin1_border
        wks['Z10'].border=thin1_border
        wks['AA10'].border=thin1_border
        wks['AB10'].border=thin1_border
        wks['AC10'].border=thin1_border

        wks.merge_cells('AG10:AM10')
        wks['AG10']=socialfund_dec.month3
        wks['AG10'].border=thin1_border
        wks['AH10'].border=thin1_border
        wks['AI10'].border=thin1_border
        wks['AJ10'].border=thin1_border
        wks['AK10'].border=thin1_border
        wks['AL10'].border=thin1_border
        wks['AM10'].border=thin1_border


        wks['J18']=socialfund_dec.hjoi3231jours1
        wks['J19']=socialfund_dec.hjos3231jours1
        wks['J20']=socialfund_dec.mi70000mois1
        wks['J21']=socialfund_dec.ms70000mois1
        wks['J22']=socialfund_dec.ms1647315mois1
        wks['J23']=socialfund_dec.salaireotal
        wks['L18']=socialfund_dec.hjoi3231jours2
        wks['L19']=socialfund_dec.hjos3231jours2
        wks['L20']=socialfund_dec.mi70000mois2
        wks['L21']=socialfund_dec.ms70000mois2
        wks['L22']=socialfund_dec.ms1647315mois2
        wks['L23']=socialfund_dec.regimetotal
        wks['P18']=socialfund_dec.hjoi3231jours3
        wks['P19']=socialfund_dec.hjos3231jours3
        wks['P20']=socialfund_dec.mi70000mois3
        wks['P21']=socialfund_dec.ms70000mois3
        wks['P22']=socialfund_dec.ms1647315mois3
        wks['P23']=socialfund_dec.travailtotal

        wks.merge_cells('V11:AC11')
        wks['V11'].border=thin_border       
        wks['W11'].border=thin_border
        wks['X11'].border=thin_border      
        wks['Y11'].border=thin_border
        wks['Z11'].border=thin_border
        wks['AA11'].border=thin_border       
        wks['AB11'].border=thin_border       
        wks['AC11'].border=thin_border

        wks.merge_cells('AF11:AM11')
        wks['AF11'].border=thin_border       
        wks['AG11'].border=thin_border
        wks['AH11'].border=thin_border      
        wks['AI11'].border=thin_border
        wks['AJ11'].border=thin_border
        wks['AK11'].border=thin_border       
        wks['AL11'].border=thin_border       
        wks['AM11'].border=thin_border

        wks['AJ15'].border=thin2_border       
        wks['AJ16'].border=thin2_border

        wks.merge_cells('A18:I18')
        wks['A18'].border=thin_border
        wks['B18'].border=thin_border
        wks['C18'].border=thin_border       
        wks['D18'].border=thin_border       
        wks['E18'].border=thin_border
        wks['F18'].border=thin_border      
        wks['G18'].border=thin_border
        wks['H18'].border=thin_border
        wks['I18'].border=thin_border 

        wks.merge_cells('J18:K18')      
        wks['J18'].border=thin_border       
        wks['K18'].border=thin_border

        wks.merge_cells('L18:O18')
        wks['L18'].border=thin_border      
        wks['M18'].border=thin_border
        wks['N18'].border=thin_border
        wks['O18'].border=thin_border

        wks.merge_cells('P18:S18')       
        wks['P18'].border=thin_border       
        wks['Q18'].border=thin_border
        wks['R18'].border=thin_border 
        wks['S18'].border=thin_border

        wks.merge_cells('T18:U18')
        wks['T18'].border=thin_border
        wks['U18'].border=thin_border

        wks.merge_cells('V18:Y18')       
        wks['V18'].border=thin_border       
        wks['W18'].border=thin_border
        wks['X18'].border=thin_border      
        wks['Y18'].border=thin_border

        wks.merge_cells('Z18:AC18')
        wks['Z18'].border=thin_border
        wks['AA18'].border=thin_border       
        wks['AB18'].border=thin_border       
        wks['AC18'].border=thin_border

        wks.merge_cells('AD18:AE18')
        wks['AD18'].border=thin_border  
        wks['AE18'].border=thin_border

        wks.merge_cells('AF18:AI18')
        wks['AF18'].border=thin_border       
        wks['AG18'].border=thin_border       
        wks['AH18'].border=thin_border
        wks['AI18'].border=thin_border 

        wks.merge_cells('AJ18:AM18')
        wks['AJ18'].border=thin_border
        wks['AK18'].border=thin_border       
        wks['AL18'].border=thin_border       
        wks['AM18'].border=thin_border


        wks.merge_cells('A19:I19')
        wks['A19'].border=thin_border
        wks['B19'].border=thin_border
        wks['C19'].border=thin_border       
        wks['D19'].border=thin_border       
        wks['E19'].border=thin_border
        wks['F19'].border=thin_border      
        wks['G19'].border=thin_border
        wks['H19'].border=thin_border
        wks['I19'].border=thin_border 

        wks.merge_cells('J19:K19')      
        wks['J19'].border=thin_border       
        wks['K19'].border=thin_border

        wks.merge_cells('L19:O19')
        wks['L19'].border=thin_border      
        wks['M19'].border=thin_border
        wks['N19'].border=thin_border
        wks['O19'].border=thin_border

        wks.merge_cells('P19:S19')       
        wks['P19'].border=thin_border       
        wks['Q19'].border=thin_border
        wks['R19'].border=thin_border 
        wks['S19'].border=thin_border

        wks.merge_cells('T19:U19')
        wks['T19'].border=thin_border
        wks['U19'].border=thin_border

        wks.merge_cells('V19:Y19')       
        wks['V19'].border=thin_border       
        wks['W19'].border=thin_border
        wks['X19'].border=thin_border      
        wks['Y19'].border=thin_border

        wks.merge_cells('Z19:AC19')
        wks['Z19'].border=thin_border
        wks['AA19'].border=thin_border       
        wks['AB19'].border=thin_border       
        wks['AC19'].border=thin_border

        wks.merge_cells('AD19:AE19')
        wks['AD19'].border=thin_border  
        wks['AE19'].border=thin_border

        wks.merge_cells('AF19:AI19')
        wks['AF19'].border=thin_border       
        wks['AG19'].border=thin_border       
        wks['AH19'].border=thin_border
        wks['AI19'].border=thin_border 

        wks.merge_cells('AJ19:AM19')
        wks['AJ19'].border=thin_border
        wks['AK19'].border=thin_border       
        wks['AL19'].border=thin_border       
        wks['AM19'].border=thin_border
       
        wks.merge_cells('A20:I20')
        wks['A20'].border=thin_border
        wks['B20'].border=thin_border
        wks['C20'].border=thin_border       
        wks['D20'].border=thin_border       
        wks['E20'].border=thin_border
        wks['F20'].border=thin_border      
        wks['G20'].border=thin_border
        wks['H20'].border=thin_border
        wks['I20'].border=thin_border 

        wks.merge_cells('J20:K20')      
        wks['J20'].border=thin_border       
        wks['K20'].border=thin_border

        wks.merge_cells('L20:O20')
        wks['L20'].border=thin_border      
        wks['M20'].border=thin_border
        wks['N20'].border=thin_border
        wks['O20'].border=thin_border

        wks.merge_cells('P20:S20')       
        wks['P20'].border=thin_border       
        wks['Q20'].border=thin_border
        wks['R20'].border=thin_border 
        wks['S20'].border=thin_border

        wks.merge_cells('T20:U20')
        wks['T20'].border=thin_border
        wks['U20'].border=thin_border

        wks.merge_cells('V20:Y20')       
        wks['V20'].border=thin_border       
        wks['W20'].border=thin_border
        wks['X20'].border=thin_border      
        wks['Y20'].border=thin_border

        wks.merge_cells('Z20:AC20')
        wks['Z20'].border=thin_border
        wks['AA20'].border=thin_border       
        wks['AB20'].border=thin_border       
        wks['AC20'].border=thin_border

        wks.merge_cells('AD20:AE20')
        wks['AD20'].border=thin_border  
        wks['AE20'].border=thin_border

        wks.merge_cells('AF20:AI20')
        wks['AF20'].border=thin_border       
        wks['AG20'].border=thin_border       
        wks['AH20'].border=thin_border
        wks['AI20'].border=thin_border 

        wks.merge_cells('AJ20:AM20')
        wks['AJ20'].border=thin_border
        wks['AK20'].border=thin_border       
        wks['AL20'].border=thin_border       
        wks['AM20'].border=thin_border

        wks.merge_cells('A21:I21')
        wks['A21'].border=thin_border
        wks['B21'].border=thin_border
        wks['C21'].border=thin_border       
        wks['D21'].border=thin_border       
        wks['E21'].border=thin_border
        wks['F21'].border=thin_border      
        wks['G21'].border=thin_border
        wks['H21'].border=thin_border
        wks['I21'].border=thin_border 

        wks.merge_cells('J21:K21')      
        wks['J21'].border=thin_border       
        wks['K21'].border=thin_border

        wks.merge_cells('L21:O21')
        wks['L21'].border=thin_border      
        wks['M21'].border=thin_border
        wks['N21'].border=thin_border
        wks['O21'].border=thin_border

        wks.merge_cells('P21:S21')       
        wks['P21'].border=thin_border       
        wks['Q21'].border=thin_border
        wks['R21'].border=thin_border 
        wks['S21'].border=thin_border

        wks.merge_cells('T21:U21')
        wks['T21'].border=thin_border
        wks['U21'].border=thin_border

        wks.merge_cells('V21:Y21')       
        wks['V21'].border=thin_border       
        wks['W21'].border=thin_border
        wks['X21'].border=thin_border      
        wks['Y21'].border=thin_border

        wks.merge_cells('Z21:AC21')
        wks['Z21'].border=thin_border
        wks['AA21'].border=thin_border       
        wks['AB21'].border=thin_border       
        wks['AC21'].border=thin_border

        wks.merge_cells('AD21:AE21')
        wks['AD21'].border=thin_border  
        wks['AE21'].border=thin_border

        wks.merge_cells('AF21:AI21')
        wks['AF21'].border=thin_border       
        wks['AG21'].border=thin_border       
        wks['AH21'].border=thin_border
        wks['AI21'].border=thin_border 

        wks.merge_cells('AJ21:AM21')
        wks['AJ21'].border=thin_border
        wks['AK21'].border=thin_border       
        wks['AL21'].border=thin_border       
        wks['AM21'].border=thin_border

        wks.merge_cells('A22:I22')
        wks['A22'].border=thin_border
        wks['B22'].border=thin_border
        wks['C22'].border=thin_border       
        wks['D22'].border=thin_border       
        wks['E22'].border=thin_border
        wks['F22'].border=thin_border      
        wks['G22'].border=thin_border
        wks['H22'].border=thin_border
        wks['I22'].border=thin_border 

        wks.merge_cells('J22:K22')      
        wks['J22'].border=thin_border       
        wks['K22'].border=thin_border

        wks.merge_cells('L22:O22')
        wks['L22'].border=thin_border      
        wks['M22'].border=thin_border
        wks['N22'].border=thin_border
        wks['O22'].border=thin_border

        wks.merge_cells('P22:S22')       
        wks['P22'].border=thin_border       
        wks['Q22'].border=thin_border
        wks['R22'].border=thin_border 
        wks['S22'].border=thin_border

        wks.merge_cells('T22:U22')
        wks['T22'].border=thin_border
        wks['U22'].border=thin_border

        wks.merge_cells('V22:Y22')       
        wks['V22'].border=thin_border       
        wks['W22'].border=thin_border
        wks['X22'].border=thin_border      
        wks['Y22'].border=thin_border

        wks.merge_cells('Z22:AC22')
        wks['Z22'].border=thin_border
        wks['AA22'].border=thin_border       
        wks['AB22'].border=thin_border       
        wks['AC22'].border=thin_border

        wks.merge_cells('AD22:AE22')
        wks['AD22'].border=thin_border  
        wks['AE22'].border=thin_border

        wks.merge_cells('AF22:AI22')
        wks['AF22'].border=thin_border       
        wks['AG22'].border=thin_border       
        wks['AH22'].border=thin_border
        wks['AI22'].border=thin_border 

        wks.merge_cells('AJ22:AM22')
        wks['AJ22'].border=thin_border
        wks['AK22'].border=thin_border       
        wks['AL22'].border=thin_border       
        wks['AM22'].border=thin_border

        wks.merge_cells('A23:I23')
        wks['A23'].border=medium_border
        wks['B23'].border=medium_border
        wks['C23'].border=medium_border       
        wks['D23'].border=medium_border       
        wks['E23'].border=medium_border
        wks['F23'].border=medium_border      
        wks['G23'].border=medium_border
        wks['H23'].border=medium_border
        wks['I23'].border=medium_border 

        wks.merge_cells('J23:K23')      
        wks['J23'].border=medium_border       
        wks['K23'].border=medium_border

        wks.merge_cells('L23:O23')
        wks['L23'].border=medium_border      
        wks['M23'].border=medium_border
        wks['N23'].border=medium_border
        wks['O23'].border=medium_border

        wks.merge_cells('P23:S23')       
        wks['P23'].border=medium_border       
        wks['Q23'].border=medium_border
        wks['R23'].border=medium_border 
        wks['S23'].border=medium_border

        wks.merge_cells('T23:U23')
        wks['T23'].border=medium_border
        wks['U23'].border=medium_border

        wks.merge_cells('V23:Y23')       
        wks['V23'].border=medium_border       
        wks['W23'].border=medium_border
        wks['X23'].border=medium_border      
        wks['Y23'].border=medium_border

        wks.merge_cells('Z23:AC23')
        wks['Z23'].border=medium_border
        wks['AA23'].border=medium_border       
        wks['AB23'].border=medium_border       
        wks['AC23'].border=medium_border

        wks.merge_cells('AD23:AE23')
        wks['AD23'].border=medium_border  
        wks['AE23'].border=medium_border

        wks.merge_cells('AF23:AI23')
        wks['AF23'].border=medium_border       
        wks['AG23'].border=medium_border       
        wks['AH23'].border=medium_border
        wks['AI23'].border=medium_border 

        wks.merge_cells('AJ23:AM23')
        wks['AJ23'].border=medium_border
        wks['AK23'].border=medium_border       
        wks['AL23'].border=medium_border       
        wks['AM23'].border=medium_border
               

        wks.merge_cells('L25:S25')
        wks['L25']=socialfund_dec.csbsactrr
        wks['M25'].border=thin_border
        wks['N25'].border=thin_border
        wks['O25'].border=thin_border
        wks['P25'].border=thin_border
        wks['Q25'].border=thin_border
        wks['R25'].border=thin_border
        wks['S25'].border=thin_border
        
        wks.merge_cells('AF25:AM25')
        wks['AF25']=socialfund_dec.csbsctrpf
        wks['AF25'].border=thin_border
        wks['AG25'].border=thin_border
        wks['AH25'].border=thin_border
        wks['AI25'].border=thin_border
        wks['AJ25'].border=thin_border
        wks['AK25'].border=thin_border
        wks['AL25'].border=thin_border
        wks['AM25'].border=thin_border

        wks.merge_cells('A30:G30')
        wks['A30'].border=thin_border
        wks['B30'].border=thin_border
        wks['C30'].border=thin_border
        wks['D30'].border=thin_border
        wks['E30'].border=thin_border
        wks['F30'].border=thin_border
        wks['G30'].border=thin_border
        
        
        wks.merge_cells('H27:I27')
        wks['H27'].border=thin_border
        wks['I27'].border=thin_border

        wks.merge_cells('J27:L27')
        wks['J27']=socialfund_dec.month1
        wks['J27'].border=thin1_border
        wks['K27'].border=thin1_border
        wks['L27'].border=thin1_border

        wks.merge_cells('H30:I30')
        wks['H29']=socialfund_dec.pcprr1
        wks['H29'].border=thin_border
        wks['I29'].border=thin_border

        wks.merge_cells('J29:L29')
        wks['J29']=socialfund_dec.pcprr2
        wks['J29'].border=thin_border
        wks['K29'].border=thin_border
        wks['L29'].border=thin_border

        wks.merge_cells('H30:I30')
        wks['H30']=socialfund_dec.ppqecp1
        wks['H30'].border=thin_border
        wks['I30'].border=thin_border

        wks.merge_cells('J30:L30')
        wks['J30']=socialfund_dec.ppqecp2
        wks['J30'].border=thin_border
        wks['K30'].border=thin_border
        wks['L30'].border=thin_border

        wks.merge_cells('M27:N27')
        wks['M27'].border=thin_border
        wks['N27'].border=thin_border

        wks.merge_cells('O27:Q27')
        wks['O27']=socialfund_dec.month4
        wks['O27'].border=thin1_border
        wks['P27'].border=thin1_border
        wks['Q27'].border=thin1_border

        wks.merge_cells('M29:N29')
        wks['M29']=socialfund_dec.pcprr11
        wks['M29'].border=thin_border
        wks['N29'].border=thin_border

        wks.merge_cells('O29:Q29')
        wks['O29']=socialfund_dec.pcprr12
        wks['O29'].border=thin_border
        wks['P29'].border=thin_border
        wks['Q29'].border=thin_border

        wks.merge_cells('M30:N30')
        wks['M30']=socialfund_dec.ppqecp11
        wks['M30'].border=thin_border
        wks['N30'].border=thin_border

        wks.merge_cells('O30:Q30')
        wks['O30']=socialfund_dec.ppqecp12
        wks['O30'].border=thin_border
        wks['P30'].border=thin_border
        wks['Q30'].border=thin_border

        wks.merge_cells('R27:S27')
        wks['R27'].border=thin_border
        wks['S27'].border=thin_border

        wks.merge_cells('T27:V27')
        wks['T27']=socialfund_dec.month5
        wks['T27'].border=thin1_border
        wks['U27'].border=thin1_border
        wks['V27'].border=thin1_border

        wks.merge_cells('T28:V28')
        wks['T28'].border=thin_border
        wks['U28'].border=thin_border
        wks['V28'].border=thin_border

        wks.merge_cells('R29:S29')
        wks['R29']=socialfund_dec.pcprr21
        wks['R29'].border=thin_border
        wks['S29'].border=thin_border
       

        wks.merge_cells('T29:V29')
        wks['T29']=socialfund_dec.pcprr22
        wks['T29'].border=thin_border
        wks['U29'].border=thin_border
        wks['V29'].border=thin_border
        
        wks.merge_cells('R30:S30')
        wks['R30']=socialfund_dec.ppqecp21
        wks['R30'].border=thin_border
        wks['S30'].border=thin_border

        wks.merge_cells('T30:V30')
        wks['T30']=socialfund_dec.ppqecp22
        wks['T30'].border=thin_border
        wks['U30'].border=thin_border
        wks['V30'].border=thin_border

        wks.merge_cells('AC28:AG28')
        wks['AC28'].border=dashed_border
        wks['AD28'].border=dashed_border
        wks['AE28'].border=dashed_border
        wks['AF28'].border=dashed_border
        wks['AG28'].border=dashed_border

        wks.merge_cells('AI28:AM28')
        wks['AI28']=socialfund_dec.date
        wks['AI28'].border=dashed_border
        wks['AJ28'].border=dashed_border
        wks['Ak28'].border=dashed_border
        wks['AL28'].border=dashed_border
        wks['AM28'].border=dashed_border
        


        wks.merge_cells('A32:G32')
        wks['A32'].border=thin_border
        wks['B32'].border=thin_border
        wks['C32'].border=thin_border
        wks['D32'].border=thin_border
        wks['E32'].border=thin_border
        wks['F32'].border=thin_border
        wks['G32'].border=thin_border

        wks.merge_cells('A33:G33')
        wks['A33'].border=thin_border
        wks['B33'].border=thin_border
        wks['C33'].border=thin_border
        wks['D33'].border=thin_border
        wks['E33'].border=thin_border
        wks['F33'].border=thin_border
        wks['G33'].border=thin_border

        wks.merge_cells('A34:G34')
        wks['A34'].border=thin_border
        wks['B34'].border=thin_border
        wks['C34'].border=thin_border
        wks['D34'].border=thin_border
        wks['E34'].border=thin_border
        wks['F34'].border=thin_border
        wks['G34'].border=thin_border

        wks.merge_cells('A35:G35')
        wks['A35'].border=thin_border
        wks['B35'].border=thin_border
        wks['C35'].border=thin_border
        wks['D35'].border=thin_border
        wks['E35'].border=thin_border
        wks['F35'].border=thin_border
        wks['G35'].border=thin_border

        wks.merge_cells('H32:L32')
        wks['H32'].border=thin_border
        wks['I32'].border=thin_border
        wks['J32'].border=thin_border
        wks['k32'].border=thin_border
        wks['L32'].border=thin_border

        wks.merge_cells('H33:L33')
        wks['H33']=socialfund_dec.pf1
        wks['H33'].border=thin_border
        wks['I33'].border=thin_border
        wks['J33'].border=thin_border
        wks['k33'].border=thin_border
        wks['L33'].border=thin_border

        wks.merge_cells('H34:L34')
        wks['H34']=socialfund_dec.at1
        wks['H34'].border=thin_border
        wks['I34'].border=thin_border
        wks['J34'].border=thin_border
        wks['k34'].border=thin_border
        wks['L34'].border=thin_border

        wks.merge_cells('H35:L35')
        wks['H35']=socialfund_dec.rr1
        wks['H35'].border=thin_border
        wks['I35'].border=thin_border
        wks['J35'].border=thin_border
        wks['k35'].border=thin_border
        wks['L35'].border=thin_border

        wks.merge_cells('M32:N32')
        wks['M32'].border=thin_border
        wks['N32'].border=thin_border

        wks.merge_cells('M33:N33')
        wks['M33']=socialfund_dec.pf2
        wks['M33'].border=thin_border
        wks['N33'].border=thin_border
        
        wks.merge_cells('M34:N34')
        wks['M34']=socialfund_dec.at2
        wks['M34'].border=thin_border
        wks['N34'].border=thin_border

        wks.merge_cells('M35:N35')
        wks['M35']=socialfund_dec.rr2
        wks['M35'].border=thin_border
        wks['N35'].border=thin_border

        wks.merge_cells('O32:T32')
        wks['O32'].border=thin_border
        wks['P32'].border=thin_border
        wks['Q32'].border=thin_border
        wks['R32'].border=thin_border
        wks['S32'].border=thin_border
        wks['T32'].border=thin_border

        wks.merge_cells('O33:T33')
        wks['O33']=socialfund_dec.pf3
        wks['O33'].border=thin_border
        wks['P33'].border=thin_border
        wks['Q33'].border=thin_border
        wks['R33'].border=thin_border
        wks['S33'].border=thin_border
        wks['T33'].border=thin_border

        wks.merge_cells('O34:T34')
        wks['O34']=socialfund_dec.at3
        wks['O34'].border=thin_border
        wks['P34'].border=thin_border
        wks['Q34'].border=thin_border
        wks['R34'].border=thin_border
        wks['S34'].border=thin_border
        wks['T34'].border=thin_border

        wks.merge_cells('O35:T35')
        wks['O35']=socialfund_dec.rr3
        wks['O35'].border=thin_border
        wks['P35'].border=thin_border
        wks['Q35'].border=thin_border
        wks['R35'].border=thin_border
        wks['S35'].border=thin_border
        wks['T35'].border=thin_border
        

        wks.merge_cells('O36:T36')
        wks['O36']=socialfund_dec.tcap
        wks['O36'].border=medium_border
        wks['P36'].border=medium_border
        wks['Q36'].border=medium_border
        wks['R36'].border=medium_border
        wks['S36'].border=medium_border
        wks['T36'].border=medium_border

        wks.merge_cells('A38:AM38')
        wks['A38'].border=thin_border
        wks['B38'].border=thin_border
        wks['C38'].border=thin_border
        wks['D38'].border=thin_border
        wks['E38'].border=thin_border
        wks['F38'].border=thin_border
        wks['G38'].border=thin_border
        wks['H38'].border=thin_border
        wks['I38'].border=thin_border
        wks['J38'].border=thin_border
        wks['K38'].border=thin_border
        wks['L38'].border=thin_border
        wks['M38'].border=thin_border
        wks['N38'].border=thin_border
        wks['O38'].border=thin_border
        wks['P38'].border=thin_border
        wks['Q38'].border=thin_border
        wks['R38'].border=thin_border
        wks['S38'].border=thin_border
        wks['T38'].border=thin_border
        wks['U38'].border=thin_border
        wks['V38'].border=thin_border
        wks['W38'].border=thin_border
        wks['X38'].border=thin_border
        wks['Y38'].border=thin_border
        wks['Z38'].border=thin_border
        wks['AA38'].border=thin_border
        wks['AB38'].border=thin_border
        wks['AC38'].border=thin_border
        wks['AD38'].border=thin_border
        wks['AE38'].border=thin_border
        wks['AF38'].border=thin_border
        wks['AG38'].border=thin_border
        wks['AH38'].border=thin_border
        wks['AI38'].border=thin_border
        wks['AJ38'].border=thin_border
        wks['AK38'].border=thin_border
        wks['AL38'].border=thin_border
        wks['AM38'].border=thin_border

        wks.merge_cells('A39:B39')
        wks['A39'].border=thin_border
        wks['B39'].border=thin_border
        wks.merge_cells('A40:B40')
        wks['A40'].border=thin_border
        wks['B40'].border=thin_border

        wks.merge_cells('C39:F39')
        wks['C39'].border=thin_border
        wks['D39'].border=thin_border
        wks['E39'].border=thin_border
        wks['F39'].border=thin_border
        wks.merge_cells('C40:F40')
        wks['C40'].border=thin_border
        wks['D40'].border=thin_border
        wks['E40'].border=thin_border
        wks['F40'].border=thin_border

        wks.merge_cells('G39:H39')
        wks['G39'].border=thin_border
        wks['H39'].border=thin_border
        wks.merge_cells('G40:H40')
        wks['G40'].border=thin_border
        wks['H40'].border=thin_border

        wks.merge_cells('I39:K39')
        wks['I39'].border=thin_border
        wks['J39'].border=thin_border
        wks['K39'].border=thin_border
        wks.merge_cells('I40:K40')
        wks['I40'].border=thin_border
        wks['J40'].border=thin_border
        wks['K40'].border=thin_border

        wks.merge_cells('L39:N39')
        wks['L39'].border=thin_border
        wks['M39'].border=thin_border
        wks['N39'].border=thin_border
        wks.merge_cells('L40:N40')
        wks['L40'].border=thin_border
        wks['M40'].border=thin_border
        wks['N40'].border=thin_border

        wks.merge_cells('O39:Q39')
        wks['O39'].border=thin_border
        wks['P39'].border=thin_border
        wks['Q39'].border=thin_border
        wks.merge_cells('O40:Q40')
        wks['O40'].border=thin_border
        wks['P40'].border=thin_border
        wks['Q40'].border=thin_border

        wks.merge_cells('R39:Y39')
        wks['R39'].border=thin_border
        wks['S39'].border=thin_border
        wks['T39'].border=thin_border
        wks['U39'].border=thin_border
        wks['V39'].border=thin_border
        wks['W39'].border=thin_border
        wks['X39'].border=thin_border
        wks['Y39'].border=thin_border
        wks.merge_cells('R40:Y40')
        wks['R40'].border=thin_border
        wks['S40'].border=thin_border
        wks['T40'].border=thin_border
        wks['U40'].border=thin_border
        wks['V40'].border=thin_border
        wks['W40'].border=thin_border
        wks['X40'].border=thin_border
        wks['Y40'].border=thin_border

        wks.merge_cells('Z39:AB39')
        wks['Z39'].border=thin_border
        wks['AA39'].border=thin_border
        wks['AB39'].border=thin_border
        wks.merge_cells('Z40:AB40')
        wks['Z40'].border=thin_border
        wks['AA40'].border=thin_border
        wks['AB40'].border=thin_border

        wks.merge_cells('AC39:AH39')
        wks['AC39'].border=thin_border
        wks['AD39'].border=thin_border
        wks['AE39'].border=thin_border
        wks['AF39'].border=thin_border
        wks['AG39'].border=thin_border
        wks['AH39'].border=thin_border
        wks.merge_cells('AC40:AH40')
        wks['AC40'].border=thin_border
        wks['AD40'].border=thin_border
        wks['AE40'].border=thin_border
        wks['AF40'].border=thin_border
        wks['AG40'].border=thin_border
        wks['AH40'].border=thin_border

        wks.merge_cells('AI39:AM39')
        wks['AI39'].border=thin_border
        wks['AJ39'].border=thin_border
        wks['AK39'].border=thin_border
        wks['AL39'].border=thin_border
        wks['AM39'].border=thin_border
        wks.merge_cells('AI40:AM40')
        wks['AI40'].border=thin_border
        wks['AJ40'].border=thin_border
        wks['AK40'].border=thin_border
        wks['AL40'].border=thin_border
        wks['AM40'].border=thin_border


        wbk.save(fl)
        fl.seek(0)
        buf=base64.encodestring(fl.read())
        ctx=dict(context)
        ctx.update({'file':buf})
        if context is None:
            context={}
        data = {}
        res= self.read(cr, uid, ids, [], context=context)
        res=  res and res[0] or {}
        data['form'] = res
        try:
                form_id= self.pool.get('ir.model.data').get_object_reference(cr, uid, 'socialfund_report_xls','socialfund_form')[1]
        except ValueError:
                form_id=False
        return{
            'type':'ir.actions.act_window',
            'view_type':'form',
            'view_mode':'form',
            'res_model':'socialfund.report.file',
            'views':[(form_id, 'form')],
            'view_id':form_id,
            'target':'new',
            'context':ctx,
        }